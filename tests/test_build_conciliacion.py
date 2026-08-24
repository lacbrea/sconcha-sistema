"""Tests de conciliacion/build_conciliacion.py: match_lote() (cruce por lote de
pago semanal de un cargo del banco contra 2+ comprobantes del mismo
proveedor, ver docstring de la funcion).

build_conciliacion.py es un script mono-modulo: corre ENTERO al importarlo
(argparse a nivel de modulo, lee el EECC, arma las hojas y guarda el .xlsx) -
no tiene un main() ni esta pensado para importarse como libreria (por eso
ningun otro archivo de tests/ lo importa directo; conciliar.py lo invoca como
subproceso). Para poder llamar match_lote() de forma aislada sin reescribir
el script:

1. Se arma un EECC sintetico minimo (1 fila de ABONO, sin cargos - alcanza
   para no disparar la guarda de "0 movimientos" de
   parsers_eecc._validar_parseo) y se importa el modulo UNA sola vez por
   archivo (fixture de modulo) con sys.argv apuntando a ese EECC, sin
   --comprobantes (asi 'comprobantes' queda en [] y CARGO_KEYS en set() al
   terminar el import - el resto del script corre igual, pero su salida, un
   .xlsx en un tmp_path descartable, no se usa ni se verifica aca).
2. Cada test reemplaza bc.comprobantes y bc.CARGO_KEYS por los datos del
   caso y llama bc.match_lote(...) directo: la MISMA funcion que usa el
   motor real contra datos reales, sin duplicar su logica en el test.

Caso real que motiva este archivo (2026-08-23, EL TEMPLO, jul-2026): el cargo
LLONTOP SANTIN JORGE LUIS de S/975.00 del 27/07 nunca cruzaba porque tenia 7
comprobantes candidatos en la ventana y el codigo viejo topaba el TAMAÑO DEL
GRUPO en 6 ("para no explotar combinatoria"), no el tamaño de la combinacion.
El fix topa el tamaño de la COMBINACION (MAX_LOTE_K, ver comentario junto a
match_lote() en build_conciliacion.py) y deja crecer el grupo.

Correr con:
    C:\\Python312\\python.exe -m pytest tests/test_build_conciliacion.py -q
"""
from __future__ import annotations

import datetime
import importlib.util
import pathlib
import sys

import openpyxl
import pytest

RAIZ_PROYECTO = pathlib.Path(__file__).resolve().parent.parent
CONC_DIR = RAIZ_PROYECTO / "conciliacion"
if str(CONC_DIR) not in sys.path:
    sys.path.insert(0, str(CONC_DIR))


def _importar_build_conciliacion(tmp_path_factory):
    """Arma un EECC sintetico minimo e importa build_conciliacion.py contra
    el (ver docstring del modulo). El formato es el 'export Excel' que
    parsers_eecc.parse_eecc_interbank_xlsx espera: hoja 'Page 1', datos desde
    la fila 13, columnas B=fecha operacion, H=cargo, I=abono, J=saldo."""
    tmp = tmp_path_factory.mktemp("build_conciliacion_import")
    eecc_path = tmp / "EC_SINTETICO.xlsx"
    salida_path = tmp / "salida_descartable.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Page 1"
    fila = 13
    ws.cell(fila, 2, "01/01/2026")   # fop
    ws.cell(fila, 3, "01/01/2026")   # fpr
    ws.cell(fila, 4, "000001")       # nro
    ws.cell(fila, 5, "DEPOSITO")     # mov
    ws.cell(fila, 6, "SALDO INICIAL DE PRUEBA")  # desc
    ws.cell(fila, 7, "")             # canal
    ws.cell(fila, 8, None)           # cargo: ninguno (no queremos entradas en CARGO_KEYS)
    ws.cell(fila, 9, 100.00)         # abono
    ws.cell(fila, 10, 100.00)        # saldo
    wb.save(eecc_path)

    sys.argv = ["build_conciliacion.py", str(eecc_path), "none", str(salida_path), "EL TEMPLO"]
    spec = importlib.util.spec_from_file_location(
        "build_conciliacion_bajo_test", CONC_DIR / "build_conciliacion.py"
    )
    modulo = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modulo)
    return modulo


@pytest.fixture(scope="module")
def bc(tmp_path_factory):
    return _importar_build_conciliacion(tmp_path_factory)


@pytest.fixture(autouse=True)
def _estado_limpio(bc):
    """match_lote() lee dos globales del modulo (comprobantes, CARGO_KEYS):
    resetearlos antes de CADA test evita que un caso contamine al siguiente
    (el modulo se importa una sola vez por archivo, no por test)."""
    bc.comprobantes = []
    bc.CARGO_KEYS = set()
    yield


def D(dia, mes, anio=2026):
    return datetime.datetime(anio, mes, dia)


def comp(total, serie, fecha_emision=None, fecha_pago=None, estado="PAGADA",
         proveedor="LLONTOP SANTIN JORGE LUIS"):
    """Fabrica un dict de comprobante con los campos que lee match_lote() (via
    esta_pagado/ref_fecha_pago/name_match): _TOTAL, _FECHA_EMISION,
    _FECHA_PAGO, _ASSIGNED, ESTADO_PAGO, PROVEEDOR. SERIE_NUMERO es solo para
    identificar el resultado en los asserts (match_lote no lo lee)."""
    return {
        "_TOTAL": total, "_FECHA_EMISION": fecha_emision, "_FECHA_PAGO": fecha_pago,
        "_CAJA_CHICA": False, "_ASSIGNED": False,
        "ESTADO_PAGO": estado, "PROVEEDOR": proveedor, "SERIE_NUMERO": serie,
    }


# -----------------------------------------------------------------------------
# Caso real: LLONTOP jul-2026, 7 candidatos (el n>6 que antes bloqueaba TODO
# el cruce por combinaciones, no solo las combinaciones caras)
# -----------------------------------------------------------------------------
def test_match_lote_llontop_jul2026_encuentra_el_subconjunto_de_7_candidatos(bc):
    """Cargo LLONTOP SANTIN JORGE LUIS, S/975.00 el 27/07/2026, descripcion
    'JORGE LUIS LLONTOP SANTIN - PESQUERO'. 7 comprobantes candidatos en la
    ventana de 15 dias (contado/PAGADA, sin FECHA_PAGO -> fallback por
    FECHA_EMISION). Antes del fix (tope en n<=6) el lote nunca se probaba por
    combinaciones porque n=7: solo se intentaba la suma completa (S/1,500),
    fallaba fuera de TOL_LOTE, y el cargo quedaba SIN COMPROBANTE. El
    subconjunto exacto 375+225+225+150=975 existe entre estos 7 candidatos."""
    bc.comprobantes = [
        comp(225.00, "F001-00001031", fecha_emision=D(12, 7)),
        comp(150.00, "F001-00001043", fecha_emision=D(15, 7)),
        comp(150.00, "F001-00001053", fecha_emision=D(17, 7)),
        comp(375.00, "F001-00001065", fecha_emision=D(19, 7)),
        comp(150.00, "F001-00001075", fecha_emision=D(21, 7)),
        comp(225.00, "F001-00001089", fecha_emision=D(25, 7)),
        comp(225.00, "F001-00001099", fecha_emision=D(26, 7)),
    ]
    resultado = bc.match_lote(975.00, D(27, 7), "JORGE LUIS LLONTOP SANTIN - PESQUERO", True)
    assert resultado is not None
    assert len(resultado) == 4
    assert round(sum(i["_TOTAL"] for i in resultado), 2) == 975.00
    # El MONTO de la combinacion ganadora es {150, 225, 225, 375} - es la unica
    # forma de sumar exactamente 975 con 4 de estos 7 montos. CUALES facturas
    # concretas de 150/225 ganan (hay 3 de cada) es ambiguo por diseño (varios
    # subconjuntos empatan en diferencia 0) y depende del orden de iteracion,
    # como documenta el docstring de match_lote(); lo que no es ambiguo es el
    # monto de cada comprobante elegido, que es lo que importa contablemente.
    assert sorted(i["_TOTAL"] for i in resultado) == [150.00, 225.00, 225.00, 375.00]


# -----------------------------------------------------------------------------
# Grupo grande sin subconjunto exacto: no debe inventar un cruce
# -----------------------------------------------------------------------------
def test_match_lote_grupo_grande_sin_subconjunto_no_cruza(bc):
    """8 comprobantes (>6, el caso que antes ni se intentaba por combinacion)
    cuya suma total es S/1,200.00. El cargo es de S/1,250.00: como supera la
    suma de TODOS los comprobantes, ningun subconjunto (de ningun tamaño)
    puede acercarse mas que el grupo completo, que ya queda a S/50.00 de
    diferencia (verificado por fuerza bruta sobre las 2**8 combinaciones) -
    muy por fuera de TOL_LOTE (+/- S/0.10). No debe cruzar nada."""
    bc.comprobantes = [
        comp(120.00, "A1", fecha_emision=D(1, 7), proveedor="PROVEEDOR SIN LOTE"),
        comp(130.00, "A2", fecha_emision=D(2, 7), proveedor="PROVEEDOR SIN LOTE"),
        comp(140.00, "A3", fecha_emision=D(3, 7), proveedor="PROVEEDOR SIN LOTE"),
        comp(160.00, "A4", fecha_emision=D(4, 7), proveedor="PROVEEDOR SIN LOTE"),
        comp(170.00, "A5", fecha_emision=D(5, 7), proveedor="PROVEEDOR SIN LOTE"),
        comp(180.00, "A6", fecha_emision=D(6, 7), proveedor="PROVEEDOR SIN LOTE"),
        comp(190.00, "A7", fecha_emision=D(7, 7), proveedor="PROVEEDOR SIN LOTE"),
        comp(110.00, "A8", fecha_emision=D(8, 7), proveedor="PROVEEDOR SIN LOTE"),
    ]
    resultado = bc.match_lote(1250.00, D(20, 7), "PROVEEDOR SIN LOTE", True)
    assert resultado is None


# -----------------------------------------------------------------------------
# Determinismo: el mismo input da siempre el mismo lote
# -----------------------------------------------------------------------------
def test_match_lote_es_determinista_mismo_input_mismo_resultado(bc):
    """Mismo escenario LLONTOP (7 candidatos, con varios subconjuntos
    empatando en la diferencia minima por los montos repetidos de 150/225):
    20 corridas independientes, cada una con una lista 'comprobantes'
    fabricada de nuevo desde cero (mismos montos, mismo orden), deben dar
    exactamente el mismo resultado. El resultado no puede depender de la
    iteracion de un dict/set sin orden garantizado entre corridas - 'groups'
    es un dict armado desde una lista con orden fijo, no un set."""
    resultados = []
    for _ in range(20):
        bc.comprobantes = [
            comp(225.00, "F1", fecha_emision=D(12, 7)),
            comp(150.00, "F2", fecha_emision=D(15, 7)),
            comp(150.00, "F3", fecha_emision=D(17, 7)),
            comp(375.00, "F4", fecha_emision=D(19, 7)),
            comp(150.00, "F5", fecha_emision=D(21, 7)),
            comp(225.00, "F6", fecha_emision=D(25, 7)),
            comp(225.00, "F7", fecha_emision=D(26, 7)),
        ]
        r = bc.match_lote(975.00, D(27, 7), "JORGE LUIS LLONTOP SANTIN - PESQUERO", True)
        resultados.append(tuple(i["SERIE_NUMERO"] for i in r) if r else None)
    assert resultados[0] is not None
    assert len(set(resultados)) == 1, f"resultados distintos entre corridas: {set(resultados)}"


# -----------------------------------------------------------------------------
# Un exacto compite con un aproximado: gana el exacto
# -----------------------------------------------------------------------------
def test_match_lote_exacto_le_gana_a_aproximado_aunque_se_evalue_despues(bc):
    """Grupo de 3 comprobantes (0.07 + 300.00 + 250.00 = 550.07). El grupo
    COMPLETO es candidato APROXIMADO (diff S/0.07, dentro de TOL_LOTE) y se
    evalua PRIMERO (match_lote prueba la suma del grupo entero antes que las
    combinaciones). El par 300.00+250.00=550.00 es EXACTO (diff 0) pero se
    evalua DESPUES, como combinacion de tamaño 2. Debe ganar el exacto: la
    funcion se queda con la MENOR diferencia de todos los candidatos, no con
    el primero que entra en tolerancia."""
    prov = "DISTRIBUCIONES ACME SAC"
    bc.comprobantes = [
        comp(0.07, "T-EXTRA", fecha_emision=D(10, 7), proveedor=prov),
        comp(300.00, "T-A", fecha_emision=D(11, 7), proveedor=prov),
        comp(250.00, "T-B", fecha_emision=D(12, 7), proveedor=prov),
    ]
    resultado = bc.match_lote(550.00, D(20, 7), "DISTRIBUCIONES ACME", True)
    assert resultado is not None
    assert sorted(i["SERIE_NUMERO"] for i in resultado) == ["T-A", "T-B"]
    assert round(sum(i["_TOTAL"] for i in resultado), 2) == 550.00
