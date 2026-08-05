#!/usr/bin/env python3
"""Herencia de Proveedor/Categoria/Tipo/Observacion desde una corrida anterior
(SCONCHA conciliacion v3.1) - modulo compartido, usado por build_conciliacion.py
via la opcion `--heredar <xlsx_anterior>`.

Caso de uso (ver SKILL.md "Heredar categorias de una corrida anterior"): hay que
regenerar el Excel de un mes (ej. porque cambio el formato de EECC usado, o se
perdieron las constancias/JSON de Gmail de esa corrida) pero SIN perder la
depuracion manual de Proveedor/Categoria que el dueño ya hizo sobre el Excel
anterior de ese mismo mes/empresa (cargos que quedaban "PENDIENTE CONSTANCIA" o
sin categoria y que el dueño corrigio a mano en el Excel).

Estrategia (clave robusta, sin depender de IDs internos que no sobreviven entre
corridas):
  - Se lee cada hoja CARGOS / "CARGOS <ALIAS>" del xlsx anterior (mismas hojas
    que genera build_conciliacion.py) y se agrupan sus filas por
    (nombre de hoja, fecha de operacion, monto del cargo redondeado a 2 decimales).
  - Si dos o mas cargos del mismo dia y monto EXACTO caen en el mismo grupo, se
    desambiguan por: (a) Nro. de operacion si el cargo nuevo lo trae y coincide
    con el de algun registro del grupo; si no, (b) orden de aparicion (FIFO): el
    archivo anterior y la corrida nueva procesan los movimientos de un mismo
    EECC exactamente en el mismo orden (fecha descendente, orden del banco
    dentro del dia - ver build_conciliacion.py), asi que consumir la cola en
    orden reproduce la asignacion correcta con altisima probabilidad.
  - Cada registro del mapa se consume una sola vez (`lookup()` lo marca `used`)
    para no asignar el mismo cargo viejo a dos cargos nuevos distintos.

Uso como modulo (ya integrado en build_conciliacion.py con --heredar):
    import heredar_categorias
    mapa = heredar_categorias.build_map('CONCILIACION EL TEMPLO - JUNIO 2026.xlsx')
    prev = heredar_categorias.lookup(mapa, 'CARGOS', '21/06/2026', 350.00, None)
    # prev = {'nro':..., 'proveedor':..., 'categoria':..., 'tipo':..., 'observacion':..., 'used': True} o None

Uso standalone (inspeccionar/volcar el mapa a JSON, util para debug o para
reusar el mapa sin releer el xlsx cada vez):
    python3 heredar_categorias.py <anterior.xlsx> <salida.json>
"""
import sys
import json
import datetime
import openpyxl

# Columnas de la hoja CARGOS (ver CARGOS_HEADERS en build_conciliacion.py),
# 0-indexado (tal como los entrega ws.iter_rows(values_only=True)):
COL_FECHA = 0
COL_NRO = 4
COL_CARGO = 8
COL_TIPO = 10
COL_CATEGORIA = 11
COL_PROVEEDOR = 12
COL_OBSERVACION = 13

# Categoria retirada por decision del usuario 2026-07-22 (ver
# reference/reglas_categorias.md): un xlsx anterior generado con la v3.1 previa
# a esa decision puede traer esta categoria; se remapea a SUNAT al heredar para
# no resucitarla.
_CATEGORIA_RETIRADA = {'RETENCION JUDICIAL': 'SUNAT'}


def _norm_fecha(v):
    """Normaliza una fecha de celda (string 'DD/MM/AAAA' o datetime) a
    'DD/MM/AAAA' de texto, para que la clave sea comparable sin importar si
    openpyxl la devuelve como texto o como datetime."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%d/%m/%Y')
    s = str(v).strip()
    return s or None


def _round2(v):
    try:
        return round(abs(float(v)), 2)
    except (TypeError, ValueError):
        return None


def _clean(v):
    if v is None:
        return ''
    return str(v).strip()


def build_map(xlsx_path):
    """Lee TODAS las hojas CARGOS / "CARGOS <ALIAS>" de `xlsx_path` y arma:
        {sheet_name: {(fecha, monto): [registro, registro, ...], ...}, ...}
    cada registro: {'nro', 'proveedor', 'categoria', 'tipo', 'observacion', 'used'}
    Las listas conservan el orden de aparicion en la hoja (= orden de la
    corrida que genero ese xlsx, fecha descendente) para el desempate FIFO de
    `lookup()`.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = {}
    for name in wb.sheetnames:
        if name != 'CARGOS' and not name.startswith('CARGOS '):
            continue
        ws = wb[name]
        bucket = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            fecha = _norm_fecha(row[COL_FECHA]) if len(row) > COL_FECHA else None
            if not fecha:
                continue  # fila TOTAL (fecha vacia) u otra fila no-cargo
            monto = _round2(row[COL_CARGO]) if len(row) > COL_CARGO else None
            if monto is None:
                continue
            nro = _clean(row[COL_NRO]) if len(row) > COL_NRO else ''
            categoria = _clean(row[COL_CATEGORIA]) if len(row) > COL_CATEGORIA else ''
            proveedor = _clean(row[COL_PROVEEDOR]) if len(row) > COL_PROVEEDOR else ''
            observacion = _clean(row[COL_OBSERVACION]) if len(row) > COL_OBSERVACION else ''
            tipo = _clean(row[COL_TIPO]) if len(row) > COL_TIPO else ''
            registro = {
                'nro': nro or None, 'proveedor': proveedor, 'categoria': categoria,
                'tipo': tipo, 'observacion': observacion, 'used': False,
            }
            bucket.setdefault((fecha, monto), []).append(registro)
        out[name] = bucket
    return out


def lookup(mapa, sheet_name, fecha, monto, nro=None):
    """Busca (y consume) el registro de la corrida anterior que corresponde al
    cargo nuevo (misma hoja, fecha+monto exactos). Si hay varios candidatos sin
    usar en el grupo, prioriza el que coincide en Nro. de operacion; si no hay
    coincidencia de numero (o el cargo nuevo no trae numero), toma el primero
    sin usar en orden de aparicion (FIFO). Devuelve None si no hay match."""
    bucket = mapa.get(sheet_name)
    if not bucket:
        return None
    fecha_n = _norm_fecha(fecha)
    monto_n = _round2(monto)
    if fecha_n is None or monto_n is None:
        return None
    candidatos = bucket.get((fecha_n, monto_n))
    if not candidatos:
        return None
    nro_n = _clean(nro) or None
    if nro_n:
        for reg in candidatos:
            if not reg['used'] and reg['nro'] == nro_n:
                reg['used'] = True
                return reg
    for reg in candidatos:
        if not reg['used']:
            reg['used'] = True
            return reg
    return None


def categoria_migrada(categoria):
    """Remapea categorias retiradas (ej. RETENCION JUDICIAL -> SUNAT, decision
    2026-07-22) para que heredar de un xlsx viejo no resucite una categoria que
    ya no existe en las reglas actuales."""
    return _CATEGORIA_RETIRADA.get(categoria, categoria)


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print('Uso: python3 heredar_categorias.py <anterior.xlsx> <salida.json>')
        sys.exit(1)
    mapa = build_map(sys.argv[1])
    # JSON no admite tuplas como clave -> se serializan como "fecha|monto"
    dump = {
        sheet: {f'{fecha}|{monto:.2f}': regs for (fecha, monto), regs in bucket.items()}
        for sheet, bucket in mapa.items()
    }
    n_total = sum(len(regs) for bucket in mapa.values() for regs in bucket.values())
    with open(sys.argv[2], 'w', encoding='utf-8') as f:
        json.dump(dump, f, ensure_ascii=False, indent=2)
    print(f'{sys.argv[1]}: {n_total} cargo(s) indexado(s) en {len(mapa)} hoja(s) -> {sys.argv[2]}')
