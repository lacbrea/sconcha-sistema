#!/usr/bin/env python3
"""Conciliacion v3.1 - EL TEMPLO / INSTITUCION CEVICHERA (SCONCHA).

Fuente de verdad del skill (ver ../SKILL.md). Parte de build_conciliacion_v2.py
(CONCILIACION\\build_conciliacion_v2.py, que queda como referencia historica) y de
la v3 (cruce de comprobantes, caja chica, VERIFICACION de conciliacion). v3.1
agrega (ver ../SKILL.md "Que hace"):

- Insumo: acepta los 3 formatos de EECC via scripts/parsers_eecc.py (formato
  detectado por contenido, no por flag): export Excel "Movimientos_..." (v3,
  ahora alternativa), PDF oficial Interbank (preferido, ancla saldo exacto),
  .xls de BBVA que en realidad es HTML (preferido, ancla saldo exacto).
- Modo multi-cuenta: varias cuentas de la MISMA empresa en un solo Excel
  (--eecc repetible o "none" + --eecc...). Cuenta principal SOLES -> hojas
  CARGOS/ABONOS de siempre. Cuentas adicionales -> hojas "CARGOS <ALIAS>" /
  "ABONOS <ALIAS>" (ALIAS = "BBVA", o "USD <ultimos 4 digitos>"). FLUJO CAJA y
  EGP consolidan SOLO cuentas en soles. Cuentas sin movimientos -> una fila en
  VERIFICACION, sin hojas propias.
- Deteccion de TRANSFERENCIA ENTRE CUENTAS de la misma empresa (mismo monto,
  fechas +/-2 dias, cargo en una cuenta = abono en otra) -> categoria
  TRANSFERENCIA ENTRE CUENTAS, tipo NO EGP, CONCILIADO=SI,
  N_COMPROBANTE="NO APLICA (TRASPASO INTERNO)".
- VERIFICACION: cuadre EXACTO por cuenta (SALDO INICIAL + ABONOS - CARGOS =
  SALDO FINAL, marca OK/ERROR) usando el saldo corrido que cada parser valida.
- --egresos <json> (2026-08, opcional): JSON intermedio armado por
  egresos_caja.py (via conciliar.py) con el reporte de egresos de caja del
  sistema de ventas (Restaurant.pe). Reemplaza, SOLO cuando se pasa este
  flag, la regla vieja de CAJA CHICA (fondo fijo S/500 + boletas del CSV de
  comprobantes) por la regla vigente desde ago-2026: rendiciones = gastos
  del reporte, reposicion = S/reposicion_semanal (del JSON, no hardcodeada)
  desde el banco. Tambien cruza los DEPOSITOS DE VENTA EN EFECTIVO del
  reporte contra los ABONOS del banco (monto exacto +/-0.05, fecha +/-1 dia,
  en cualquier cuenta cargada) y marca el abono que cruza en su propia hoja
  ABONOS. Sin --egresos, el comportamiento es EXACTAMENTE el de antes (ver
  FONDO_CAJA_CHICA mas abajo).

Todo lo de v2/v3 se conserva: reglas de categorias/proveedores, TIPO para
EGP/flujo de caja, hojas CARGOS/ABONOS/FLUJO CAJA/EGP/CAJA CHICA/VERIFICACION/EECC,
cruce de comprobantes, ITF resuelto antes que constancias, verificacion de cuadre.

Uso (compatible con v3, un solo EECC):
  python3 build_conciliacion.py <eecc> <constancias.json|none> <salida.xlsx> "<EMPRESA>" \\
      [--comprobantes facturas.csv] [--pendientes pendientes.json]

Uso multi-cuenta (v3.1, varios EECC de la misma empresa en un solo Excel):
  python3 build_conciliacion.py none <constancias.json|none> <salida.xlsx> "<EMPRESA>" \\
      --eecc EC_4134.pdf --eecc EC_BBVA_8579.xls --eecc EC_4388.pdf \\
      [--comprobantes facturas.csv] [--pendientes pendientes.json]

  <eecc> puede ser "none" si todas las cuentas se pasan por --eecc (repetible);
  si <eecc> NO es "none" se agrega a la lista junto con los --eecc (el primero
  con movimientos en SOLES pasa a ser la cuenta principal, sin importar el
  orden en que se hayan listado).

  EMPRESA = "EL TEMPLO" | "INSTITUCION CEVICHERA" (se usa en titulos y en la logica
  de empresas hermanas / caja chica por local).
"""
import openpyxl, json, re, datetime, sys, os, csv, argparse
from itertools import combinations
from collections import Counter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import parsers_eecc
import heredar_categorias

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
ap = argparse.ArgumentParser(description='Conciliacion bancaria SCONCHA v3.1')
ap.add_argument('bank_eecc', help='EECC principal (xlsx|pdf|xls-HTML BBVA), o "none" si se usa solo --eecc')
ap.add_argument('constancias', help='cons_<cuenta>.json o "none"')
ap.add_argument('salida_xlsx')
ap.add_argument('empresa', help='"EL TEMPLO" o "INSTITUCION CEVICHERA"')
ap.add_argument('--comprobantes', default=None, help='CSV export del Sheet "SCONCHA - Facturas"')
ap.add_argument('--pendientes', default=None, help='ruta de salida para pendientes.json')
ap.add_argument('--eecc', action='append', default=[], metavar='ARCHIVO',
                 help='EECC adicional de la MISMA empresa (repetible) - modo multi-cuenta')
ap.add_argument('--banco', default=None, choices=[None, 'interbank', 'bbva'],
                 help='(obsoleto v3.1: el formato se detecta solo; se ignora si se pasa)')
ap.add_argument('--heredar', default=None, metavar='XLSX_ANTERIOR',
                 help='Excel de una corrida anterior de la MISMA empresa/mes (ver scripts/heredar_categorias.py): '
                      'hereda Proveedor/Categoria/Tipo/Observacion de los cargos ya depurados a mano cuando esta '
                      'corrida no los determina automaticamente (ITF/COMISION/JUDICIAL/traspasos siempre ganan la '
                      'regla nueva). Util para regenerar un mes (ej. cambio de formato de EECC, o se perdio el '
                      'JSON de constancias de Gmail) sin perder la depuracion previa.')
ap.add_argument('--pdf-password', default=None, metavar='CONTRASENA',
                 help='Contrasena para abrir los EECC en PDF que vengan cifrados (jul-2026: Interbank empezo a '
                      'proteger los PDF de "Cuenta Negocio" con el RUC del titular). Se aplica a TODOS los PDF '
                      'de esta corrida (bank_eecc + --eecc); si alguno no esta cifrado, se ignora para ese '
                      'archivo sin error.')
ap.add_argument('--egresos', default=None, metavar='JSON',
                 help='JSON intermedio (armado por egresos_caja.py via conciliar.py) con el reporte de egresos '
                      'de caja del sistema de ventas: {"gastos": [...], "depositos": [...], '
                      '"reposicion_semanal": N}. Si se pasa, la hoja CAJA CHICA usa la regla vigente desde '
                      'ago-2026 (rendiciones = gastos del reporte, reposicion semanal desde banco) en vez de la '
                      'regla vieja de fondo fijo S/500, y se cruzan los DEPOSITOS DE VENTA EN EFECTIVO del '
                      'reporte contra los ABONOS del banco. Sin este flag, el comportamiento es el de siempre.')
args = ap.parse_args()

if args.banco:
    print(f'NOTA: --banco {args.banco} ya no es necesario (el formato se detecta por contenido); se ignora.')

eecc_paths = list(args.eecc)
if args.bank_eecc.lower() != 'none':
    eecc_paths = [args.bank_eecc] + eecc_paths
if not eecc_paths:
    ap.error('No se paso ningun EECC (ni posicional ni --eecc).')

cj, dst, EMP = args.constancias, args.salida_xlsx, args.empresa.upper()
SISTER_CAT = 'INSTITUCION CEVICHERA' if 'TEMPLO' in EMP else 'EL TEMPLO'
EMP_KEY = 'TEMPLO' if 'TEMPLO' in EMP else 'CEVICHERA'

CAJA_CHICA_LOCAL = {
    'TEMPLO': ('LINCE', 'YESSICA (ROJAS COBEÑAS)'),
    'CEVICHERA': ('MIRAFLORES', 'JORDANO / YORDANO (ALVAREZ)'),
}
FONDO_CAJA_CHICA = 500.00

MESES_ES = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO',
            'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

cons = json.load(open(cj)) if cj.lower() != 'none' and os.path.exists(cj) else []

# EGRESOS_DATA es None cuando no se paso --egresos: ese es el interruptor que
# decide toda la rama vieja vs nueva de CAJA CHICA (ver esa seccion). El JSON
# lo arma conciliar.py (gastos/depositos ya separados por egresos_caja.py +
# reposicion_semanal leido de config.yaml, nunca hardcodeado aca).
EGRESOS_DATA = json.load(open(args.egresos, encoding='utf-8')) if args.egresos else None

HEREDAR_MAP = heredar_categorias.build_map(args.heredar) if args.heredar else None
HEREDAR_STATS = {'prov': 0, 'cat': 0, 'reclasificaciones': []}

DIAS = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']


def N(s):
    return re.sub(r'[^A-Z0-9 ]', ' ', str(s or '').upper()).strip()


ACCENTS = str.maketrans('ÁÉÍÓÚáéíóúÑñÜü', 'AEIOUaeiouNnUu')


def strip_accents(s):
    return str(s or '').translate(ACCENTS)


def norm(s):
    """Normalizacion para matching laxo de nombres: mayusculas, sin tildes, solo A-Z0-9."""
    return N(strip_accents(s))


STOPWORDS = {'SAC', 'S', 'A', 'C', 'SA', 'SRL', 'EIRL', 'E', 'I', 'R', 'L', 'DE', 'DEL',
             'LA', 'EL', 'LOS', 'LAS', 'Y', 'CIA'}

# Tolerancia de redondeo banco vs comprobante en el cruce 1-a-1 (caso real
# jun-2026: factura APUDEX S/2,348.76 vs cargo S/2,348.75). Siempre comparar
# con round(...,2): la resta cruda de floats puede dar 0.010000000000218
# y romper un <= 0.01. Dif >= 0.01 se anota en OBSERVACION.
TOL_INDIVIDUAL = 0.05

# Tolerancia del cruce por lote: la suma de un lote semanal puede diferir del
# cargo por redondeos acumulados de varios comprobantes (jun-2026: lotes
# MOGOFRAN y LLONTOP con diferencias de hasta S/0.06 que solo se conciliaban
# a mano). Dif >= 0.01 se anota en OBSERVACION.
TOL_LOTE = 0.10

# Tokens que aparecen en razones sociales de medio Peru: compartirlos NO
# significa que sea el mismo proveedor (jun-2026: cargo MAPFRE cruzo con
# "GRUPO GIOBRE PERU S.A.C." solo por "PERU").
GENERIC_TOKENS = {'PERU', 'PERUANA', 'PERUANO', 'PERUANAS', 'PERUANOS', 'GRUPO',
                  'INVERSIONES', 'SERVICIOS', 'SERVICIO', 'GENERALES', 'GENERAL',
                  'CORPORACION', 'DISTRIBUIDORA', 'DISTRIBUCIONES', 'COMERCIAL',
                  'COMERCIALIZADORA', 'NEGOCIOS', 'EMPRESA', 'EMPRESAS', 'COMPANIA',
                  'INDUSTRIA', 'INDUSTRIAS', 'INDUSTRIAL', 'IMPORTACIONES',
                  'EXPORTACIONES', 'SOLUCIONES', 'REPRESENTACIONES', 'INTERNACIONAL',
                  'NACIONAL', 'MULTISERVICIOS', 'CONSULTORES', 'ASOCIADOS', 'HERMANOS',
                  'SOCIEDAD', 'ANONIMA', 'CERRADA', 'LIMITADA', 'LIMA'}


def name_tokens(s):
    # El banco pega el nombre del beneficiario al numero de operacion sin
    # espacio ("...MOGOFRAN3140991"): separar letras de digitos antes de
    # tokenizar, y descartar tokens puramente numericos (un nro de operacion
    # no es evidencia de nombre).
    t = re.sub(r'(?<=[A-Z])(?=[0-9])|(?<=[0-9])(?=[A-Z])', ' ', norm(s))
    return set(x for x in t.split() if x not in STOPWORDS and len(x) >= 3 and not x.isdigit())


def name_match(a, b):
    """Match laxo: los tokens genericos no cuentan como evidencia por si solos.
    Cruza si hay un token significativo compartido de >=4 letras, o uno de 3
    letras respaldado por al menos otro token compartido (aunque sea generico):
    'SERVICIOS GENERALES NEA S.A.C.' cruza por NEA+SERVICIOS+GENERALES, pero un
    cargo MAPFRE no cruza con 'GRUPO GIOBRE PERU S.A.C.' solo por PERU (jun-2026)."""
    ta, tb = name_tokens(a), name_tokens(b)
    if not ta or not tb:
        return False
    common = ta & tb
    signif = common - GENERIC_TOKENS
    if any(len(t) >= 4 for t in signif):
        return True
    return bool(signif) and len(common) >= 2


def pdate(s):
    try:
        return datetime.datetime.strptime(str(s).strip(), '%d/%m/%Y')
    except Exception:
        return None


def pdate_flex(s):
    """Parser de fecha tolerante para el CSV de comprobantes (dd/mm/aaaa o aaaa-mm-dd)."""
    s = str(s or '').strip()
    if not s:
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(s, fmt)
        except Exception:
            continue
    return None


def dia(s):
    d = pdate(s)
    return DIAS[d.weekday()] if d else ''


def semana(s):
    d = pdate(s)
    if not d:
        return ''
    return f'S{min((d.day - 1) // 7 + 1, 5)}'


def _semanas_en_periodo(ini, fin):
    """Cuenta cuantos 'buckets' S1..S5 (mismo criterio que semana(), bandas de
    7 dias del mes calendario) toca el periodo [ini, fin]. Se usa SOLO en la
    hoja CAJA CHICA con --egresos, para calcular la cadencia esperada de
    reposicion_semanal sin inventar una nocion de 'semana' nueva (reutiliza
    la misma que ya usa FLUJO CAJA)."""
    if not ini or not fin:
        return 0
    semanas = set()
    d = ini
    un_dia = datetime.timedelta(days=1)
    while d <= fin:
        semanas.add(min((d.day - 1) // 7 + 1, 5))
        d += un_dia
    return len(semanas)


cidx = {}
for c in cons:
    cidx.setdefault((c['fecha'], round(c['monto'], 2)), []).append(c['para'])

PROV_RULES = [('CASTILLO JUSTO', 'ALQUILER'), ('PEYON', 'PRESTAMO CP'), ('PALOMINO', 'GASTOS ADMIN'),
              ('ULTRAFRIO', 'PESQUERO'), ('PC10266592', 'PRESTAMO'), ('BONAVISTA', 'BAR'),
              ('JORDANO', 'CAJA CHICA'), ('KAMBISTA', 'ALQUILER'), ('VASQUEZ ORMENO', 'ILLAWARA'),
              ('ASOCIACION PERUANA DE AUTORES', 'GASTOS ADMIN'), ('AUTORES', 'GASTOS ADMIN'), ('RESTAURANT', 'SISTEMA'),
              ('SPOTIFY', 'GASTOS ADMIN'), ('BARJA', 'DESCARTABLES'), ('VELARDE', 'PLANILLA'), ('BRICE', 'PLANILLA'),
              ('MATTA', 'CONTABILIDAD'), ('COBE', 'CAJA CHICA'), ('BARRIADA', 'BARRIADA'), ('PIZARRO', 'VERDURAS'),
              ('AQUATEC', 'BAR'), ('PEFARER', 'LIMPIEZA'), ('MOGOFRAN', 'GAS'), ('MOGOLLON CASTRO', 'GAS'),
              ('ZAPATA', 'ABARROTES'), ('RICCE', 'ABARROTES'), ('LLONTOP', 'PESQUERO'), ('WILTON', 'ABARROTES'),
              ('DOLMER', 'BAR'), ('MAPFRE', 'SEGUROS'), ('PISCO', 'BAR'), ('HUAMANI', 'BAR'),
              ('MULTICOSAS', 'MENAJE / EQUIPAMIENTO'), ('EKAMA', 'BAR'), ('MAKRO', 'PESQUERO'), ('FOOD RETAIL', 'ABARROTES'),
              ('PLAZA VEA', 'ABARROTES'), ('APUDEX', 'ABARROTES'), ('HORTIFRUT', 'BAR'), ('SODIMAC', 'MANTENIMIENTO'),
              ('GASTROTEC', 'BAR'), ('CONTROLTEC', 'MANTENIMIENTO'), ('CONECTA RETAIL', 'MANTENIMIENTO'),
              ('CENCOSUD', 'LIMPIEZA'), ('METRO', 'LIMPIEZA'), ('PUNTO GRAFICO', 'MKT'), ('ROJAS ANAYA', 'MKT'),
              ('PACHECO', 'MANTENIMIENTO'), ('INKAFARMA', 'ABARROTES'), ('TERMINAL PESQUERO', 'PESQUERO'),
              ('MERCADO CAQUETA', 'VERDURAS'), ('DELIVERY HERO', 'MKT'), ('CASTRO BREA', 'PRESTAMO'),
              ('T CONTADORES', 'CONTABILIDAD'), ('TCONTADORES', 'CONTABILIDAD'), ('FREZCO', 'PESQUERO'),
              ('DIRECTV', 'GASTOS ADMIN'), ('INSUMA', 'ABARROTES'), ('PROHIGIENE', 'LIMPIEZA')]


def cat_by_prov(p):
    d = N(p)
    if 'BARRIADA' in d: return 'BARRIADA'                     # antes que CEVICHERA
    if 'INSTITUCION' in d or 'CEVICHERA' in d: return 'INSTITUCION CEVICHERA'
    if 'EL TEMPLO' in d: return 'EL TEMPLO'
    for k, c in PROV_RULES:
        if k in d: return c
    return ''


def cat_by_desc(mov, desc):
    # OJO: 'mov' (columna corta "Movimiento") solo existe en el export xlsx; en
    # los parsers PDF/BBVA todo va en 'desc' y mov queda ''. Por eso ITF/COMISION
    # se buscan en 'a' (mov+desc combinados), no solo en 'v' (mov) - si no, un
    # PDF Interbank deja los ~80 renglones de ITF por mes sin categoria bancaria
    # (cae en la heuristica de "persona" -> PROPINA por ser un monto chico).
    d = N(desc); v = N(mov); a = d + ' ' + v
    if 'ITF' in a: return 'GASTOS ADMIN'
    if 'JUDICIAL' in a: return 'SUNAT'                         # RET.JUDICIAL (BBVA) o JUDICIALES (Interbank) = embargo SUNAT (decision usuario 2026-07-22)
    if 'COMISION' in a or 'COMIS' in a or 'COMI ' in a: return 'GASTOS ADMIN'
    if 'PLANILLA' in a or 'CTS' in a: return 'PLANILLA'
    if 'PROPINA' in a: return 'PROPINA'
    if 'LUZ DEL SUR' in d: return 'LUZ DEL SUR'
    if 'SEDAPAL' in d: return 'SEDAPAL'
    if 'MAPFRE' in d: return 'SEGUROS'
    if 'BARRIADA' in d: return 'BARRIADA'
    if 'CEVICHERA' in d or 'INSTITUCION' in d: return 'INSTITUCION CEVICHERA'
    if 'EL TEMPLO' in d: return 'EL TEMPLO'
    if 'ON EMPRESAS' in d or 'INTERNET' in d: return 'INTERNET'
    if 'RENTA' in d or 'ESSALUD' in d or 'ONP' in d: return 'SUNAT'
    if 'SNT NPS' in a or 'PAG SNT' in a: return 'SUNAT'
    if 'CRED COMERC' in d: return 'PRESTAMO'
    return cat_by_prov(desc)


def person_cat(desc, amount, fecha):
    d = N(desc)
    if 'I BANC' in d or d in ('', '-'): return ''
    dt = pdate(fecha)
    if dt:
        if amount < 200: return 'PROPINA'
        if amount >= 200 and (dt.day in (14, 15, 16) or dt.day >= 28): return 'PLANILLA'
    return ''


TIPO = {
    'PESQUERO': 'COSTO DE VENTAS', 'ABARROTES': 'COSTO DE VENTAS', 'VERDURAS': 'COSTO DE VENTAS',
    'GAS': 'COSTO DE VENTAS', 'BAR': 'COSTO DE VENTAS', 'DESCARTABLES': 'COSTO DE VENTAS',
    'PLANILLA': 'PERSONAL', 'PROPINA': 'PERSONAL',
    'LUZ DEL SUR': 'SERVICIOS Y OPERACION', 'SEDAPAL': 'SERVICIOS Y OPERACION', 'INTERNET': 'SERVICIOS Y OPERACION',
    'ALQUILER': 'SERVICIOS Y OPERACION', 'SEGUROS': 'SERVICIOS Y OPERACION', 'MANTENIMIENTO': 'SERVICIOS Y OPERACION',
    'LIMPIEZA': 'SERVICIOS Y OPERACION', 'MENAJE / EQUIPAMIENTO': 'SERVICIOS Y OPERACION', 'MKT': 'SERVICIOS Y OPERACION',
    'GASTOS ADMIN': 'SERVICIOS Y OPERACION', 'CONTABILIDAD': 'SERVICIOS Y OPERACION', 'SISTEMA': 'SERVICIOS Y OPERACION',
    'BARRIADA': 'SERVICIOS Y OPERACION', 'FACTURAS': 'SERVICIOS Y OPERACION',
    'SUNAT': 'IMPUESTOS',
    'PRESTAMO': 'FINANCIAMIENTO (NO EGP)', 'PRESTAMO CP': 'FINANCIAMIENTO (NO EGP)',
    'INSTITUCION CEVICHERA': 'TRANSFERENCIA ENTRE EMPRESAS (NO EGP)', 'EL TEMPLO': 'TRANSFERENCIA ENTRE EMPRESAS (NO EGP)',
    'TRANSFERENCIA ENTRE CUENTAS': 'TRASPASO ENTRE CUENTAS (NO EGP)',
    'CAJA CHICA': 'POR RENDIR', 'ILLAWARA': 'POR RENDIR',
    'PENDIENTE CONSTANCIA': 'PENDIENTE', '': 'PENDIENTE', 'SIN CATEGORIA': 'PENDIENTE',
}


def tipo_of(cat):
    return TIPO.get(cat, 'SERVICIOS Y OPERACION')


# ---------------------------------------------------------------------------
# Carga de EECC (uno o varios) + eleccion de cuenta principal + alias
# ---------------------------------------------------------------------------
def elegir_principal(cuentas):
    """La cuenta principal es la primera en SOLES con movimientos (sin importar
    el orden en que se hayan pasado los --eecc). Si ninguna cumple (todas USD o
    todas sin movimientos), cae a la primera de la lista tal cual."""
    for i, c in enumerate(cuentas):
        if c['meta']['moneda'] == 'PEN' and c['meta']['n_movimientos'] > 0:
            return i
    return 0


def alias_de(meta):
    if meta['banco'] == 'BBVA':
        base = 'BBVA'
    else:
        last4 = re.sub(r'\D', '', meta.get('cuenta') or '')[-4:] or '????'
        base = f'USD {last4}' if meta['moneda'] == 'USD' else last4
    return base


cuentas = []
for path in eecc_paths:
    movs, meta = parsers_eecc.parse_eecc(path, password=args.pdf_password)
    cuentas.append({'path': path, 'movs': movs, 'meta': meta, 'alias': None})

principal_idx = elegir_principal(cuentas)
cuentas[principal_idx]['alias'] = None  # sin sufijo: hojas CARGOS/ABONOS de siempre
usados = {}
for i, c in enumerate(cuentas):
    if i == principal_idx:
        continue
    a = alias_de(c['meta'])
    if a in usados:
        usados[a] += 1
        a = f'{a} ({usados[a]})'
    else:
        usados[a] = 1
    c['alias'] = a

MULTI_CUENTA = len(cuentas) > 1
data = cuentas[principal_idx]['movs']  # cuenta principal (compatibilidad con el resto del script)
meta_principal = cuentas[principal_idx]['meta']

# (monto, fecha) de TODOS los cargos de todas las cuentas: se usa para no
# dejar que un cargo con dd>0 le "robe" un comprobante/lote a otro cargo del
# mismo monto cuya fecha coincide exacta con la fecha de pago (caso LLONTOP
# jun-2026: dos cargos de S/615, el lote con FECHA_PAGO 15/06 es del cargo
# del 15/06, no del 16/06 que se procesa primero por orden descendente).
CARGO_KEYS = set()
for _c in cuentas:
    for _m in _c['movs']:
        if _m['cargo'] not in (None, ''):
            _dt = pdate(_m['fop'])
            if _dt is not None:
                CARGO_KEYS.add((round(abs(float(_m['cargo'])), 2), _dt))

# Periodo cubierto por el EECC principal (para titulos dinamicos y alerta de caja chica).
_all_dates = [pdate(d['fop']) for d in data if pdate(d['fop'])]
periodo_ini = min(_all_dates) if _all_dates else None
periodo_fin = max(_all_dates) if _all_dates else None
if meta_principal.get('periodo_mes'):
    PERIODO_LABEL = f"{MESES_ES[meta_principal['periodo_mes'] - 1]} {meta_principal['periodo_anio']}"
elif periodo_fin:
    _mesyr = Counter((d.month, d.year) for d in _all_dates)
    _m, _y = _mesyr.most_common(1)[0][0]
    PERIODO_LABEL = f'{MESES_ES[_m - 1]} {_y}'
else:
    PERIODO_LABEL = ''

# ---------------------------------------------------------------------------
# Deteccion de TRANSFERENCIA ENTRE CUENTAS (solo tiene sentido con >1 cuenta)
# ---------------------------------------------------------------------------
def detectar_transferencias(cuentas):
    """Cargo en una cuenta de la empresa = abono en otra cuenta de la misma
    empresa, mismo monto (+/- 1 centimo), fechas +/-2 dias, misma moneda.
    Caso real (jun-2026, EL TEMPLO): BBVA -4,200.00 (30/05) = abono Interbank
    4134 'TRAN TIL Transf EL TE' +4,200.00 (30/05); igual para 300, 2900, 1200,
    1000. Match voraz (greedy) por menor diferencia de dias primero."""
    ids = set()
    hallados = []
    if len(cuentas) < 2:
        return ids, hallados

    def es_candidato(desc):
        d = N(desc)
        # ya se resuelven por otra regla (traspaso a empresa hermana) -> no
        # reinterpretar como transferencia ENTRE CUENTAS propias:
        if any(k in d for k in ('CEVICHERA', 'INSTITUCION', 'EL TEMPLO')):
            return False
        # depositos de efectivo/cheque: no son transferencia electronica entre
        # cuentas propias, y su monto+fecha puede coincidir por azar con un
        # traspaso real (visto en los datos de prueba: un traspaso a la
        # hermana de S/1,000 coincidio en fecha con un DEPOS. EN CTA de BBVA
        # de S/1,000 - false positive que este filtro evita).
        if any(k in d for k in ('DEPOS EN CTA', 'DEP EFECTIVO', 'INGRESO EN EFECTIVO')):
            return False
        return True

    cargos = []
    abonos = []
    for ci, c in enumerate(cuentas):
        for m in c['movs']:
            dt = pdate(m['fop'])
            if dt is None:
                continue
            if not es_candidato(m['desc']):
                continue
            if m['cargo']:
                cargos.append((ci, m, dt, round(m['cargo'], 2)))
            if m['abono']:
                abonos.append((ci, m, dt, round(m['abono'], 2)))
    candidatos = []
    for xi, (cai, cm, cdt, camt) in enumerate(cargos):
        for yi, (bai, bm, bdt, bamt) in enumerate(abonos):
            if cai == bai:
                continue
            if cuentas[cai]['meta']['moneda'] != cuentas[bai]['meta']['moneda']:
                continue
            if abs(camt - bamt) > 0.01:
                continue
            dd = abs((cdt - bdt).days)
            if dd > 2:
                continue
            candidatos.append((dd, xi, yi))
    candidatos.sort(key=lambda t: t[0])
    used_c, used_a = set(), set()
    for dd, xi, yi in candidatos:
        if xi in used_c or yi in used_a:
            continue
        used_c.add(xi); used_a.add(yi)
        cai, cm, cdt, camt = cargos[xi]
        bai, bm, bdt, bamt = abonos[yi]
        ids.add(cm['_id']); ids.add(bm['_id'])
        hallados.append({
            'monto': camt,
            'cuenta_cargo': cuentas[cai]['alias'] or 'PRINCIPAL', 'fecha_cargo': cm['fop'], 'desc_cargo': cm['desc'],
            'cuenta_abono': cuentas[bai]['alias'] or 'PRINCIPAL', 'fecha_abono': bm['fop'], 'desc_abono': bm['desc'],
        })
    return ids, hallados


transfer_ids, transferencias = detectar_transferencias(cuentas)

# ---------------------------------------------------------------------------
# Cruce de DEPOSITOS DE VENTA EN EFECTIVO (reporte --egresos) vs ABONOS del
# banco. Regla del dueno (2026-08): el reporte de egresos de caja trae los
# depositos que la caja envia al banco; cada uno deberia aparecer como un
# abono real en el EECC. El cruce NO se restringe a la redaccion que use el
# banco (BBVA describe "DEPOS. EN CTA." o "INGRESO EN EFECTIVO") - cualquier
# abono que calce en monto+fecha es candidato, para no depender de un texto
# que puede cambiar de banco a banco o de mes a mes.
# ---------------------------------------------------------------------------
TOL_DEPOSITO = 0.05


def cruzar_depositos_caja_chica(cuentas, depositos):
    """Para cada deposito del reporte busca UN abono (monto exacto +/-0.05,
    fecha +/-1 dia) en CUALQUIERA de las cuentas cargadas, sin reutilizar un
    abono ya asignado a otro deposito (mismo espiritu que '_ASSIGNED' en el
    cruce de comprobantes: dos depositos del mismo monto en fechas cercanas
    no deben pelearse por el mismo abono en silencio - el de menor
    diferencia de dias/monto gana, procesando los depositos en orden de
    fecha para que el resultado sea determinista).

    Devuelve (match: {abono['_id']: deposito_dict}, depositos_sin_match: [dict]).
    """
    abonos = []
    for c in cuentas:
        for m in c['movs']:
            if not m['abono']:
                continue
            dt = pdate(m['fop'])
            if dt is None:
                continue
            abonos.append((m, dt, round(float(m['abono']), 2)))

    usados = set()
    match = {}
    sin_match = []
    orden = sorted(depositos, key=lambda d: pdate(d.get('fecha')) or datetime.datetime.min)
    for dep in orden:
        dfecha = pdate(dep.get('fecha'))
        try:
            dmonto = round(float(dep.get('monto')), 2)
        except (TypeError, ValueError):
            sin_match.append(dep)
            continue
        if dfecha is None:
            sin_match.append(dep)
            continue
        best_idx = None; best_key = None
        for idx, (m, dt, amt) in enumerate(abonos):
            if idx in usados:
                continue
            diff_amt = round(abs(amt - dmonto), 2)
            if diff_amt > TOL_DEPOSITO:
                continue
            dd = abs((dt - dfecha).days)
            if dd > 1:
                continue
            key = (diff_amt, dd)
            if best_key is None or key < best_key:
                best_idx, best_key = idx, key
        if best_idx is None:
            sin_match.append(dep)
        else:
            usados.add(best_idx)
            match[abonos[best_idx][0]['_id']] = dep
    return match, sin_match


def abonos_deposito_sin_cruzar(cuentas, deposito_match):
    """Abonos que el banco describe como deposito de efectivo (DEPOS. EN
    CTA. / INGRESO EN EFECTIVO) pero que ningun deposito del reporte reclamo
    en cruzar_depositos_caja_chica() -el caso contrario al anterior-, para
    no dejarlos pasar en silencio tampoco."""
    encontrados = []
    for c in cuentas:
        for m in c['movs']:
            if not m['abono'] or m['_id'] in deposito_match:
                continue
            dN = N(m['desc'])
            if ('DEPOS' in dN and 'CTA' in dN) or 'INGRESO EN EFECTIVO' in dN:
                encontrados.append(m)
    return encontrados


DEPOSITO_MATCH, DEPOSITOS_SIN_MATCH = (
    cruzar_depositos_caja_chica(cuentas, EGRESOS_DATA.get('depositos', [])) if EGRESOS_DATA else ({}, [])
)
ABONOS_DEPOSITO_SIN_MATCH = abonos_deposito_sin_cruzar(cuentas, DEPOSITO_MATCH) if EGRESOS_DATA else []

# ---------------------------------------------------------------------------
# Comprobantes (Sheet "SCONCHA - Facturas" exportado a CSV)
# ---------------------------------------------------------------------------
COMPROBANTES_CARGADOS = bool(args.comprobantes)
comprobantes = []
if args.comprobantes:
    with open(args.comprobantes, encoding='utf-8-sig', newline='') as f:
        for raw in csv.DictReader(f):
            row = {k.strip().upper(): (v.strip() if isinstance(v, str) else v) for k, v in raw.items()}
            if EMP_KEY not in norm(row.get('EMPRESA', '')):
                continue

            def fnum(x):
                x = str(x or '').replace(',', '').strip()
                try:
                    return float(x)
                except Exception:
                    return None

            row['_TOTAL'] = fnum(row.get('TOTAL'))
            row['_FECHA_EMISION'] = pdate_flex(row.get('FECHA_EMISION'))
            row['_FECHA_PAGO'] = pdate_flex(row.get('FECHA_PAGO'))
            row['_CAJA_CHICA'] = norm(row.get('CAJA_CHICA', '')) == 'SI'
            row['_ASSIGNED'] = False
            if row['_TOTAL'] is None:
                continue
            comprobantes.append(row)


def ref_fecha_pago(c):
    """Fecha de referencia para cruzar un comprobante contra un cargo del banco:
    FECHA_PAGO si existe; si es CONTADO sin fecha de pago registrada, FECHA_EMISION."""
    return c['_FECHA_PAGO'] or c['_FECHA_EMISION']


def esta_pagado(c):
    """Solo los comprobantes ya PAGADOS pueden cruzar contra un cargo real del banco;
    uno 'pendiente' o 'programada' todavia no genero el cargo (evita falsos positivos
    por coincidencia de monto+fecha con una factura que aun no se paga)."""
    return norm(c.get('ESTADO_PAGO', '')) == 'PAGADA'


def match_individual(monto, fecha_cargo, prov_hint, require_name):
    """Cruce 1-a-1: mismo TOTAL (+/- S/0.05, redondeado a 2 decimales) y fecha de
    referencia dentro de +/-3 dias; prefiere el de menor diferencia de monto y
    luego de dias. Desambiguacion de fecha exacta: si dd>0 y existe otro cargo
    del mismo monto exactamente en la fecha de referencia del comprobante (ver
    CARGO_KEYS), el comprobante se reserva para ese cargo y no se ofrece aqui."""
    best = None; bestkey = None
    for c in comprobantes:
        if c['_ASSIGNED'] or not esta_pagado(c):
            continue
        ref = ref_fecha_pago(c)
        if ref is None or fecha_cargo is None:
            continue
        diff_amt = round(abs(c['_TOTAL'] - monto), 2)
        if diff_amt > TOL_INDIVIDUAL:
            continue
        dd = abs((ref - fecha_cargo).days)
        if dd > 3:
            continue
        if dd > 0 and (round(c['_TOTAL'], 2), ref) in CARGO_KEYS:
            continue  # otro cargo de fecha exacta esta esperando este comprobante
        if require_name and prov_hint and not name_match(c.get('PROVEEDOR', ''), prov_hint):
            continue
        key = (diff_amt, dd)
        if bestkey is None or key < bestkey:
            best, bestkey = c, key
    return best


def match_lote(monto, fecha_cargo, prov_hint, require_name):
    """Cruce por lote de pago semanal: mismo proveedor + misma FECHA_PAGO, suma de 2+
    comprobantes cuya diferencia con el monto del cargo esta dentro de +/- S/0.10
    (TOL_LOTE, redondeado a 2 decimales - los lotes acumulan redondeos de varios
    comprobantes). Prueba primero el grupo completo y luego combinaciones (grupos
    pequeños, hasta 6 comprobantes por proveedor/fecha, para no explotar
    combinatoria); de todos los grupos/combos candidatos dentro de tolerancia,
    gana el de MENOR diferencia (un lote exacto siempre le gana a uno aproximado).

    Fallback sin FECHA_PAGO: si el comprobante no trae FECHA_PAGO (CSV incompleto),
    se agrupa por proveedor + emision <= fecha del cargo, hasta 15 dias antes (un
    solo pago de lote puede cubrir facturas emitidas en varios dias).

    Desambiguacion de fecha exacta: si un grupo (por FECHA_PAGO real) no es de la
    fecha del cargo actual y existe otro cargo del mismo monto exactamente en esa
    fecha (ver CARGO_KEYS), el grupo no se le ofrece a este cargo - se reserva
    para el cargo de fecha exacta."""
    if fecha_cargo is None:
        return None
    groups = {}
    for c in comprobantes:
        if c['_ASSIGNED'] or not esta_pagado(c):
            continue
        if require_name and prov_hint and not name_match(c.get('PROVEEDOR', ''), prov_hint):
            continue
        fp = c['_FECHA_PAGO']
        if fp is not None:
            if abs((fp - fecha_cargo).days) > 3:
                continue
            key = (norm(c.get('PROVEEDOR', '')), fp)
        else:
            # Fallback sin FECHA_PAGO: lote semanal pagado hasta 15 dias despues
            # de emitidas las facturas (emision <= fecha del cargo). Se agrupa por
            # proveedor (un solo pago cubre emisiones de varios dias).
            fe = c['_FECHA_EMISION']
            if fe is None:
                continue
            dd_em = (fecha_cargo - fe).days
            if dd_em < 0 or dd_em > 15:
                continue
            key = (norm(c.get('PROVEEDOR', '')), 'EMISION')
        groups.setdefault(key, []).append(c)
    best = None; best_diff = None
    for key, items in groups.items():
        if len(items) < 2:
            continue
        # Si la FECHA_PAGO del grupo no es la del cargo actual y existe OTRO
        # cargo del mismo monto exactamente en esa fecha, este grupo es de ese
        # otro cargo (caso LLONTOP jun-2026: dos cargos de S/615, el lote con
        # FECHA_PAGO 15/06 le corresponde al cargo del 15/06, no al del 16/06
        # que se procesa antes por orden descendente). El fallback EMISION no
        # lleva este guard (no tiene una fecha "exacta" que defender).
        if key[1] != 'EMISION' and key[1] != fecha_cargo and (round(monto, 2), key[1]) in CARGO_KEYS:
            continue
        cand = []
        total_all = round(sum(i['_TOTAL'] for i in items), 2)
        cand.append((round(abs(total_all - monto), 2), items))
        n = len(items)
        if n <= 6:
            for r in range(2, n):
                for combo in combinations(items, r):
                    d = round(abs(round(sum(i['_TOTAL'] for i in combo), 2) - monto), 2)
                    cand.append((d, list(combo)))
        for d, grp in cand:
            if d > TOL_LOTE:
                continue
            if best_diff is None or d < best_diff:
                best, best_diff = grp, d
    return best


thin = Side(style='thin', color='BFBFBF'); border = Border(thin, thin, thin, thin)
hfill = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF', size=10)
tfill = PatternFill('solid', fgColor='D9E1F2'); tfont = Font(bold=True, size=10)
catfill = PatternFill('solid', fgColor='E2EFDA'); profill = PatternFill('solid', fgColor='DDEBF7')
pendfill = PatternFill('solid', fgColor='FCE4D6')
okfill = PatternFill('solid', fgColor='C6EFCE'); badfill = PatternFill('solid', fgColor='FFC7CE')
trapfill = PatternFill('solid', fgColor='E4DFEC')


def style_header(sh, headers, widths):
    for i, h in enumerate(headers, 1):
        c = sh.cell(1, i, h); c.font = hfont; c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, w in enumerate(widths, 1): sh.column_dimensions[get_column_letter(i)].width = w
    sh.freeze_panes = 'A2'


wb = openpyxl.Workbook()
wb.remove(wb.active)

CARGOS_HEADERS = ['Fecha de operacion', 'Dia', 'Semana', 'Fecha de proceso', 'Nro. de operacion', 'Movimiento', 'Descripcion', 'Canal',
                  'Cargo (S/)', 'Saldo contable', 'Tipo', 'Categoria', 'Proveedor', 'Observacion',
                  'N Comprobante', 'Link comprobante', 'Conciliado', 'Motivo pendiente']
CARGOS_WIDTHS = [15, 11, 8, 15, 15, 22, 28, 9, 12, 13, 26, 20, 32, 28, 28, 40, 12, 24]
ABONOS_HEADERS = ['Fecha de operacion', 'Dia', 'Semana', 'Fecha de proceso', 'Nro. de operacion', 'Movimiento', 'Descripcion', 'Canal',
                   'Abono (S/)', 'Saldo contable', 'Categoria', 'Cliente/Origen', 'Observacion']
ABONOS_WIDTHS = [15, 11, 8, 15, 15, 22, 28, 9, 12, 13, 20, 32, 28]


def procesar_cuenta_cargos(sheet_name, movs, transfer_ids, heredar_map=None, heredar_stats=None):
    """Construye la hoja CARGOS (o "CARGOS <ALIAS>") para una cuenta. Devuelve un
    dict con los acumulados de esa cuenta (para consolidar FLUJO CAJA/EGP/CAJA
    CHICA/VERIFICACION en el llamador)."""
    sc = wb.create_sheet(sheet_name)
    style_header(sc, CARGOS_HEADERS, CARGOS_WIDTHS)
    row = 2; total = 0; kc = {}; explic = 0; pend = []; cat_tot = {}; flujo = {}
    benef_amt = 0; sinbenef_amt = 0
    n_conc_si = 0; motivo_counts = {}
    caja_chica_repuesto_rows = []
    n_cargos_data = sum(1 for d in movs if d['cargo'] not in (None, ''))
    for d in sorted(movs, key=lambda x: (pdate(x['fop']) or datetime.datetime.min), reverse=True):
        if d['cargo'] in (None, ''): continue
        va = abs(round(float(d['cargo']), 2)); amt = -va; total += va
        prov = ''; obs = ''; mN = N(d['mov']); dN = N(d['desc'])
        # is_ibanc: 'I BANC' exacto en el campo Movimiento del xlsx, O canal
        # I-BANC (parsers PDF/BBVA: "N/D I-BANC + COM.O/C" queda con canal
        # 'I-BANC' y esa frase completa en Descripcion, no coincide con el
        # chequeo de igualdad exacta dN=='I BANC' que asumia el xlsx).
        is_ibanc = ('I BANC' in mN) or (dN == 'I BANC') or (d.get('canal') == 'I-BANC')
        key = (str(d['fop']).strip(), va)
        es_transferencia_entre_cuentas = d.get('_id') in transfer_ids
        if es_transferencia_entre_cuentas:
            prov = 'TRANSFERENCIA ENTRE CUENTAS'; obs = 'TRASPASO INTERNO ENTRE CUENTAS DE LA MISMA EMPRESA'
        elif 'ITF' in mN or 'ITF' in dN:
            prov = 'ITF'; obs = 'ITF'
        elif key in cidx:
            names = sorted(set(cidx[key]))
            if len(names) == 1: prov = names[0]; obs = 'CONSTANCIA INTERBANK'
            else:
                i = kc.get(key, 0); prov = names[i % len(names)]; kc[key] = i + 1
                obs = 'REVISAR - FECHA/MONTO COMPARTIDO'
        elif is_ibanc: prov = 'PENDIENTE CONSTANCIA'; obs = 'FALTA CONSTANCIA GMAIL'
        elif 'TRANSFER' in mN and dN not in ('', '-', 'I BANC'): prov = str(d['desc']).strip(); obs = 'BENEFICIARIO DEL ESTADO'
        elif dN not in ('', '-', 'I BANC'): prov = str(d['desc']).strip(); obs = 'DESCRIPCION DEL ESTADO'

        # ---- Herencia de Proveedor (--heredar), ANTES de calcular Categoria ----
        # Se hace antes que cat_by_prov/person_cat a proposito: si esta corrida no
        # identifico al beneficiario (I-BANC sin constancia -> PENDIENTE CONSTANCIA,
        # tipico cuando no hay JSON de Gmail), pero el archivo anterior SI lo tenia
        # (constancia que ya no esta disponible), recuperar ese Proveedor permite que
        # cat_by_prov() lo reconozca de nuevo por su regla normal (ej. "VELARDE" ->
        # PLANILLA) en vez de caer en la heuristica generica person_cat() por monto.
        # Los traspasos entre cuentas nunca heredan (la regla nueva siempre gana).
        prev = None
        if heredar_map is not None and not es_transferencia_entre_cuentas:
            prev = heredar_categorias.lookup(heredar_map, sheet_name, d['fop'], va, d.get('nro'))
            if prev and prev['proveedor'] and ((not prov) or prov == 'PENDIENTE CONSTANCIA'):
                prov = prev['proveedor']
                obs = prev['observacion'] or 'HEREDADO DE CORRIDA ANTERIOR (SIN CONSTANCIA ESTA VEZ)'
                if heredar_stats is not None: heredar_stats['prov'] += 1

        if es_transferencia_entre_cuentas:
            cat = 'TRANSFERENCIA ENTRE CUENTAS'
        else:
            cat_regla = cat_by_desc(d['mov'], d['desc'])
            if not cat_regla and prov and prov != 'PENDIENTE CONSTANCIA': cat_regla = cat_by_prov(prov)
            if cat_regla:
                cat = cat_regla
            else:
                cat = person_cat(prov or d['desc'], va, d['fop'])
                if not cat and prov == 'PENDIENTE CONSTANCIA': cat = 'PENDIENTE CONSTANCIA'
            # ---- Herencia de Categoria (--heredar) ----
            # "cat_regla" es lo que determinaron las reglas de verdad (descripcion del
            # banco o proveedor conocido, incluye JUDICIAL->SUNAT, ITF, etc.) - esas
            # SIEMPRE ganan, nunca se pisan. Lo que NO viene de una regla (blanco,
            # PENDIENTE CONSTANCIA, o solo el heuristico "monto chico=PROPINA" de
            # person_cat, que es una aproximacion, no una regla real - ver
            # reference/reglas_categorias.md seccion 3) SI se puede heredar.
            if prev and prev['categoria']:
                old_cat = heredar_categorias.categoria_migrada(prev['categoria'])
                if not cat_regla and old_cat:
                    cat = old_cat
                    if heredar_stats is not None: heredar_stats['cat'] += 1
                elif cat_regla and old_cat and old_cat != cat_regla:
                    # la corrida nueva SI determino categoria por una regla real, pero
                    # difiere de la del archivo anterior -> reclasificacion real, se
                    # reporta (la regla nueva gana, no se pisa).
                    if heredar_stats is not None:
                        heredar_stats['reclasificaciones'].append({
                            'sheet': sheet_name, 'fecha': d['fop'], 'monto': va,
                            'proveedor': prov or prev['proveedor'],
                            'categoria_anterior': old_cat, 'categoria_nueva': cat_regla,
                        })
        tp = tipo_of(cat)
        if cat and cat != 'PENDIENTE CONSTANCIA': explic += va
        else: pend.append((d['fop'], d['mov'], str(d['desc'])[:24], va))
        if prov == 'PENDIENTE CONSTANCIA': sinbenef_amt += va
        else: benef_amt += va
        cat_tot[cat or 'SIN CATEGORIA'] = cat_tot.get(cat or 'SIN CATEGORIA', 0) + va
        sm = semana(d['fop']); flujo.setdefault(sm, {}).setdefault(cat or 'SIN CATEGORIA', 0)
        flujo[sm][cat or 'SIN CATEGORIA'] += va

        # ---- Conciliacion contra comprobantes ----
        n_comprobante = ''; link_comprobante = ''; conciliado = 'NO'; motivo = ''
        # (mismo cuidado que en cat_by_desc: mirar mN y dN, no solo mN - en
        # PDF/BBVA todo el texto vive en dN, mN siempre es '')
        mN_fee = any('ITF' in x or 'COMISION' in x or 'COMIS' in x or 'COMI ' in x or 'PORTE' in x for x in (mN, dN))
        fecha_cargo_dt = pdate(d['fop'])
        if es_transferencia_entre_cuentas:
            conciliado = 'SI'; n_comprobante = 'NO APLICA (TRASPASO INTERNO)'
        elif mN_fee:
            conciliado = 'SI'; n_comprobante = 'NO APLICA'
        elif cat == SISTER_CAT:
            conciliado = 'SI'; n_comprobante = 'NO APLICA (TRASPASO)'
        elif cat == 'CAJA CHICA':
            conciliado = 'SI'; n_comprobante = 'NO APLICA (CAJA CHICA)'
            caja_chica_repuesto_rows.append((d['fop'], va))
        elif COMPROBANTES_CARGADOS:
            require_name = prov not in ('', 'PENDIENTE CONSTANCIA')
            prov_hint = prov if require_name else ''
            m = match_individual(va, fecha_cargo_dt, prov_hint, require_name)
            if m:
                m['_ASSIGNED'] = True
                conciliado = 'SI'
                n_comprobante = m.get('SERIE_NUMERO', '')
                link_comprobante = m.get('LINK_DRIVE', '')
                dif = round(abs(m['_TOTAL'] - va), 2)
                if dif >= 0.01:
                    obs = (obs + ' | ' if obs else '') + f'DIF S/{dif:.2f} VS COMPROBANTE (TOLERANCIA REDONDEO)'
            else:
                lote = match_lote(va, fecha_cargo_dt, prov_hint, require_name)
                if lote:
                    for i in lote: i['_ASSIGNED'] = True
                    conciliado = 'SI'
                    n_comprobante = ' + '.join(i.get('SERIE_NUMERO', '') for i in lote)
                    link_comprobante = ' '.join(i.get('LINK_DRIVE', '') for i in lote if i.get('LINK_DRIVE'))
                    dif_lote = round(abs(round(sum(i['_TOTAL'] for i in lote), 2) - va), 2)
                    if dif_lote >= 0.01:
                        obs = (obs + ' | ' if obs else '') + f'DIF S/{dif_lote:.2f} VS SUMA DE LOTE (TOLERANCIA REDONDEO)'
                    if any(i['_FECHA_PAGO'] is None for i in lote):
                        obs = (obs + ' | ' if obs else '') + 'LOTE POR FECHA EMISION (SIN FECHA_PAGO)'
                else:
                    motivo = 'PENDIENTE CONSTANCIA' if prov == 'PENDIENTE CONSTANCIA' else 'SIN COMPROBANTE'
        else:
            motivo = 'SIN CONSTANCIA' if prov == 'PENDIENTE CONSTANCIA' else 'SIN COMPROBANTE'

        if conciliado == 'SI': n_conc_si += 1
        else: motivo_counts[motivo] = motivo_counts.get(motivo, 0) + 1

        vals = [d['fop'], dia(d['fop']), sm, d['fpr'], d['nro'], d['mov'], d['desc'], d['canal'], amt, d['saldo']]
        for i, v in enumerate(vals, 1):
            c = sc.cell(row, i, v); c.border = border; c.font = Font(size=9); c.alignment = Alignment(vertical='center')
            if i in (9, 10): c.number_format = '#,##0.00'
        sc.cell(row, 11, tp).border = border; sc.cell(row, 11).font = Font(size=8)
        cc = sc.cell(row, 12, cat); cc.border = border; cc.font = Font(size=9, bold=bool(cat and cat != 'PENDIENTE CONSTANCIA')); cc.fill = catfill
        pc = sc.cell(row, 13, prov); pc.border = border; pc.font = Font(size=9)
        if es_transferencia_entre_cuentas: pc.fill = trapfill; cc.fill = trapfill
        elif prov == 'PENDIENTE CONSTANCIA': pc.fill = pendfill; cc.fill = pendfill
        elif prov: pc.fill = profill
        oc = sc.cell(row, 14, obs); oc.border = border; oc.font = Font(size=8); oc.alignment = Alignment(vertical='center', wrap_text=True)
        ncc = sc.cell(row, 15, n_comprobante); ncc.border = border; ncc.font = Font(size=8)
        lcc = sc.cell(row, 16, link_comprobante); lcc.border = border; lcc.font = Font(size=7)
        coc = sc.cell(row, 17, conciliado); coc.border = border; coc.font = Font(size=9, bold=True)
        coc.fill = okfill if conciliado == 'SI' else badfill
        mpc = sc.cell(row, 18, motivo); mpc.border = border; mpc.font = Font(size=8); mpc.alignment = Alignment(vertical='center', wrap_text=True)
        if motivo: mpc.fill = pendfill
        row += 1
    sc.cell(row, 8, 'TOTAL').font = tfont; sc.cell(row, 8).fill = tfill; sc.cell(row, 8).border = border; sc.cell(row, 8).alignment = Alignment(horizontal='right')
    tc = sc.cell(row, 9, -round(total, 2)); tc.font = tfont; tc.fill = tfill; tc.number_format = '#,##0.00'; tc.border = border
    for col in range(10, 19): sc.cell(row, col).fill = tfill; sc.cell(row, col).border = border
    sc_last = row - 1
    sc.auto_filter.ref = f'A1:R{sc_last}'
    return {
        'sheet': sc, 'last_row': sc_last, 'total': total, 'cat_tot': cat_tot, 'flujo': flujo,
        'caja_chica_rows': caja_chica_repuesto_rows, 'n_conc_si': n_conc_si, 'n_cargos_data': n_cargos_data,
        'motivo_counts': motivo_counts, 'pend': pend, 'explic': explic, 'benef_amt': benef_amt,
        'sinbenef_amt': sinbenef_amt,
    }


def procesar_cuenta_abonos(sheet_name, movs, transfer_ids, deposito_match=None):
    """deposito_match ({abono['_id']: deposito_dict}, ver
    cruzar_depositos_caja_chica) es None/{} en el modo sin --egresos: en ese
    caso 'dep' siempre da falsy abajo y la columna Observacion queda
    EXACTAMENTE como antes - es lo que garantiza que el modo viejo no cambia
    ni un byte de esta hoja. Se reutiliza la columna Observacion existente en
    vez de agregar una nueva (menos invasivo: mismo criterio que ya usa esa
    columna en CARGOS para varias notas distintas) - transferencia entre
    cuentas y deposito de caja chica no deberian coincidir nunca en el mismo
    abono (detectar_transferencias ya excluye depositos/ingresos en efectivo
    de sus candidatos), pero por las dudas la transferencia gana la prioridad."""
    sa = wb.create_sheet(sheet_name)
    style_header(sa, ABONOS_HEADERS, ABONOS_WIDTHS)
    deposito_match = deposito_match or {}
    row = 2; totA = 0
    for d in sorted(movs, key=lambda x: (pdate(x['fop']) or datetime.datetime.min), reverse=True):
        if d['abono'] in (None, ''): continue
        totA += abs(float(d['abono']))
        es_transferencia = d.get('_id') in transfer_ids
        dep = deposito_match.get(d.get('_id'))
        vals = [d['fop'], dia(d['fop']), semana(d['fop']), d['fpr'], d['nro'], d['mov'], d['desc'], d['canal'], d['abono'], d['saldo']]
        for i, v in enumerate(vals, 1):
            c = sa.cell(row, i, v); c.border = border; c.font = Font(size=9); c.alignment = Alignment(vertical='center')
            if i in (9, 10): c.number_format = '#,##0.00'
        cat_cell = sa.cell(row, 11, 'TRANSFERENCIA ENTRE CUENTAS' if es_transferencia else '')
        if es_transferencia:
            obs_txt = 'TRASPASO INTERNO ENTRE CUENTAS DE LA MISMA EMPRESA'
        elif dep:
            obs_txt = f"CONCILIADO CAJA CHICA (DEPOSITO DE VENTA {dep.get('fecha', '')} S/{float(dep.get('monto') or 0):.2f} - REPORTE EGRESOS)"
        else:
            obs_txt = ''
        obs_cell = sa.cell(row, 13, obs_txt)
        for col in (11, 12, 13): sa.cell(row, col).border = border
        if es_transferencia: cat_cell.fill = trapfill
        elif dep: obs_cell.fill = okfill
        row += 1
    sa.cell(row, 8, 'TOTAL').font = tfont; sa.cell(row, 8).fill = tfill; sa.cell(row, 8).border = border; sa.cell(row, 8).alignment = Alignment(horizontal='right')
    ta = sa.cell(row, 9, round(totA, 2)); ta.font = tfont; ta.fill = tfill; ta.number_format = '#,##0.00'; ta.border = border
    sa_last = row - 1
    sa.auto_filter.ref = f'A1:M{sa_last}'
    return {'sheet': sa, 'last_row': sa_last, 'total': totA, 'n_abonos': sum(1 for d in movs if d['abono'] not in (None, ''))}


def construir_eecc_sheet(sheet_name, movs, meta):
    """Hoja EECC de auditoria: los movimientos ya normalizados (independiente
    del formato de origen). Reemplaza el 'copiar la hoja original' de v3 (que
    solo existia para el export xlsx) por algo uniforme para los 3 formatos."""
    se2 = wb.create_sheet(sheet_name)
    hdr = ['Fecha operacion', 'Fecha proceso', 'Nro operacion', 'Descripcion', 'Canal', 'Cargo', 'Abono', 'Saldo contable']
    style_header(se2, hdr, [14, 14, 14, 44, 12, 12, 12, 14])
    se2.cell(1, 1).value = se2.cell(1, 1).value  # no-op, mantiene estilo
    info = (f"{meta.get('banco', '')} {meta.get('formato', '')} | cuenta {meta.get('cuenta', '')} | "
            f"{meta.get('moneda', '')} | saldo inicial {meta.get('saldo_inicial')} | "
            f"saldo final {meta.get('saldo_final')} | ancla exacto: {'SI' if meta.get('anchor_ok') else 'NO'}")
    r = 2
    for d in movs:
        vals = [d['fop'], d['fpr'], d['nro'], d['desc'], d['canal'], d['cargo'], d['abono'], d['saldo']]
        for i, v in enumerate(vals, 1):
            c = se2.cell(r, i, v); c.border = border; c.font = Font(size=9)
            if i in (6, 7, 8) and v not in (None, ''): c.number_format = '#,##0.00'
        r += 1
    se2.freeze_panes = 'A2'
    return info


# ---------------------------------------------------------------------------
# Procesar todas las cuentas (principal primero, luego las adicionales)
# ---------------------------------------------------------------------------
resultados = []  # uno por cuenta CON movimientos
sheet_aliases = []
for c in cuentas:
    alias = c['alias']
    sheet_cargos = 'CARGOS' if alias is None else f'CARGOS {alias}'
    sheet_abonos = 'ABONOS' if alias is None else f'ABONOS {alias}'
    sheet_eecc = 'EECC' if alias is None else f'EECC {alias}'
    if c['meta']['n_movimientos'] == 0:
        resultados.append({'cuenta': c, 'vacia': True})
        continue
    rc = procesar_cuenta_cargos(sheet_cargos, c['movs'], transfer_ids, HEREDAR_MAP, HEREDAR_STATS)
    ra = procesar_cuenta_abonos(sheet_abonos, c['movs'], transfer_ids, DEPOSITO_MATCH)
    eecc_info = construir_eecc_sheet(sheet_eecc, c['movs'], c['meta'])
    resultados.append({'cuenta': c, 'vacia': False, 'cargos': rc, 'abonos': ra, 'eecc_info': eecc_info,
                        'sheet_cargos': sheet_cargos, 'sheet_abonos': sheet_abonos})
    sheet_aliases.append(alias)

principal_res = next(r for r in resultados if not r.get('vacia') and r['cuenta'] is cuentas[principal_idx])

# Consolidacion FLUJO CAJA / EGP: SOLO cuentas en SOLES (ver SKILL.md - la USD
# se lista con sus propias hojas CARGOS/ABONOS pero no se consolida).
cat_tot = {}; flujo = {}; total = 0; totA = 0
caja_chica_repuesto_rows = []
n_conc_si = 0; n_cargos_data = 0; motivo_counts = {}
n_cargos = 0; n_abonos = 0
for r in resultados:
    if r.get('vacia'):
        continue
    es_pen = r['cuenta']['meta']['moneda'] == 'PEN'
    rc, ra = r['cargos'], r['abonos']
    n_conc_si += rc['n_conc_si']
    n_cargos_data += rc['n_cargos_data']
    for k, v in rc['motivo_counts'].items():
        motivo_counts[k] = motivo_counts.get(k, 0) + v
    if es_pen:
        n_cargos += rc['n_cargos_data']
        n_abonos += ra['n_abonos']
        total += rc['total']; totA += ra['total']
        for k, v in rc['cat_tot'].items(): cat_tot[k] = cat_tot.get(k, 0) + v
        for sm, d2 in rc['flujo'].items():
            flujo.setdefault(sm, {})
            for k, v in d2.items(): flujo[sm][k] = flujo[sm].get(k, 0) + v
        caja_chica_repuesto_rows.extend(rc['caja_chica_rows'])

explic = sum(v for k, v in cat_tot.items() if k != 'SIN CATEGORIA' and k != 'PENDIENTE CONSTANCIA')
pend_display = principal_res['cargos']['pend']  # detalle de "sin categoria" impreso al final: cuenta principal

# ---------------------------------------------------------------------------
# FLUJO CAJA
# ---------------------------------------------------------------------------
sf = wb.create_sheet('FLUJO CAJA')
semanas = ['S1', 'S2', 'S3', 'S4', 'S5']
cats = sorted(cat_tot.keys(), key=lambda c: (tipo_of(c), c))
hdr = ['Tipo', 'Categoria'] + semanas + ['TOTAL MES']
style_header(sf, hdr, [30, 22, 11, 11, 11, 11, 11, 14])
row = 2
for c in cats:
    sf.cell(row, 1, tipo_of(c)).font = Font(size=9); sf.cell(row, 2, c).font = Font(size=9, bold=True)
    rt = 0
    for j, s in enumerate(semanas, 3):
        v = flujo.get(s, {}).get(c, 0); rt += v
        cell = sf.cell(row, j, round(v, 2) if v else None); cell.number_format = '#,##0.00'; cell.font = Font(size=9)
    tcell = sf.cell(row, 8, round(rt, 2)); tcell.number_format = '#,##0.00'; tcell.font = Font(size=9, bold=True); tcell.fill = tfill
    for cc2 in range(1, 9): sf.cell(row, cc2).border = border
    row += 1
sf.cell(row, 2, 'TOTAL EGRESOS').font = tfont; sf.cell(row, 2).fill = tfill
for j, s in enumerate(semanas, 3):
    v = sum(flujo.get(s, {}).values()); cell = sf.cell(row, j, round(v, 2)); cell.number_format = '#,##0.00'; cell.font = tfont; cell.fill = tfill
sf.cell(row, 8, round(total, 2)).number_format = '#,##0.00'; sf.cell(row, 8).font = tfont; sf.cell(row, 8).fill = tfill
for cc2 in range(1, 9): sf.cell(row, cc2).border = border
if MULTI_CUENTA:
    row += 2
    nota = sf.cell(row, 1, 'Consolida solo cuentas en SOLES. Cuentas USD se listan en sus propias hojas CARGOS/ABONOS pero no se suman aqui.')
    nota.font = Font(size=8, italic=True)

# ---------------------------------------------------------------------------
# EGP
# ---------------------------------------------------------------------------
se = wb.create_sheet('EGP')
se['A1'] = f'ESTADO DE GANANCIAS Y PERDIDAS (PRELIMINAR) - {EMP} - {PERIODO_LABEL}'; se['A1'].font = Font(bold=True, size=12)
se.column_dimensions['A'].width = 34; se.column_dimensions['B'].width = 16; se.column_dimensions['C'].width = 44
r = 3


def wr(a, b, bold=False, fill=None, note=''):
    global r
    ca = se.cell(r, 1, a); cb = se.cell(r, 2, b); cc2 = se.cell(r, 3, note)
    ca.font = Font(bold=bold, size=10); cb.font = Font(bold=bold, size=10); cc2.font = Font(size=8, italic=True)
    if b not in (None, ''): cb.number_format = '#,##0.00'
    if fill: ca.fill = cb.fill = fill
    r += 1


wr('INGRESOS (abonos del mes)*', round(totA, 2), True, tfill,
   '*incluye ventas y traspasos; depurar en hoja ABONOS' + (' - solo cuentas SOLES' if MULTI_CUENTA else ''))
r += 1
by_tipo = {}
for c, v in cat_tot.items(): by_tipo.setdefault(tipo_of(c), []).append((c, v))
op_total = 0
for tp in ['COSTO DE VENTAS', 'PERSONAL', 'SERVICIOS Y OPERACION', 'IMPUESTOS']:
    if tp not in by_tipo: continue
    st = sum(v for _, v in by_tipo[tp]); op_total += st
    wr(tp, round(st, 2), True, PatternFill('solid', fgColor='F2F2F2'))
    for c, v in sorted(by_tipo[tp]): wr('   ' + c, round(v, 2))
r += 1
wr('TOTAL EGRESOS OPERATIVOS', round(op_total, 2), True, tfill)
wr('RESULTADO OPERATIVO (aprox)', round(totA - op_total, 2), True, PatternFill('solid', fgColor='E2EFDA'),
   'Ingresos - egresos operativos. NO incluye financiamiento/traspasos/pendientes')
r += 1
wr('PARTIDAS FUERA DEL EGP', '', True)
for tp in ['FINANCIAMIENTO (NO EGP)', 'TRANSFERENCIA ENTRE EMPRESAS (NO EGP)', 'TRASPASO ENTRE CUENTAS (NO EGP)', 'POR RENDIR', 'PENDIENTE']:
    if tp not in by_tipo: continue
    st = sum(v for _, v in by_tipo[tp])
    wr(tp, round(st, 2), False, PatternFill('solid', fgColor='FCE4D6') if 'PEND' in tp else None)
    for c, v in sorted(by_tipo[tp]): wr('   ' + c, round(v, 2))

# ---------------------------------------------------------------------------
# CAJA CHICA
# ---------------------------------------------------------------------------
scc = wb.create_sheet('CAJA CHICA')
scc.column_dimensions['A'].width = 16; scc.column_dimensions['B'].width = 30
scc.column_dimensions['C'].width = 16; scc.column_dimensions['D'].width = 16
scc.column_dimensions['E'].width = 46
rc2 = 1
pendientes_caja_chica = []  # para pendientes.json (tema conciliacion)


def wcc(a='', b='', c='', d='', note='', bold=False, fill=None, fmt_cols=()):
    global rc2
    vals = [a, b, c, d]
    for i, v in enumerate(vals, 1):
        cell = scc.cell(rc2, i, v)
        cell.font = Font(bold=bold, size=10)
        if i in fmt_cols and v not in (None, ''): cell.number_format = '#,##0.00'
        if fill: cell.fill = fill
    ncell = scc.cell(rc2, 5, note); ncell.font = Font(size=8, italic=True)
    rc2 += 1


local_key = EMP_KEY
local_nombre, local_resp = CAJA_CHICA_LOCAL.get(local_key, ('', ''))
wcc(f'CAJA CHICA - {EMP} - {PERIODO_LABEL}', '', '', '', '', True, tfill)

if EGRESOS_DATA:
    # -----------------------------------------------------------------------
    # Regla vigente desde ago-2026 (decision del dueno, ver --egresos mas
    # arriba): reposicion semanal desde el banco, sin fondo fijo. Rendiciones
    # = gastos del reporte de egresos de caja (Restaurant.pe), no boletas del
    # CSV de comprobantes (ese CSV puede seguir usandose para otras cosas del
    # cruce, pero ya no es la fuente de la caja chica).
    # -----------------------------------------------------------------------
    reposicion_semanal = float(EGRESOS_DATA.get('reposicion_semanal') or 0)
    wcc('Local', local_nombre, 'Responsable', local_resp,
        'Regla vigente desde ago-2026: reposicion semanal desde el banco (sin fondo fijo, ver config.yaml). '
        'Rendiciones = gastos del reporte de egresos de caja, no boletas del CSV.', True)
    rc2 += 1
    wcc('REPOSICION SEMANAL (config)', round(reposicion_semanal, 2), '', '',
        'conciliacion.empresas[].caja_chica.reposicion_semanal en config.yaml - nunca hardcodeada en el motor',
        True, tfill, fmt_cols=(2,))
    rc2 += 1

    gastos_reporte = EGRESOS_DATA.get('gastos') or []
    gastos_periodo = []
    for g in gastos_reporte:
        gf = pdate(g.get('fecha'))
        if periodo_ini and periodo_fin and gf and not (periodo_ini <= gf <= periodo_fin):
            continue
        gastos_periodo.append(g)
    gastos_periodo.sort(key=lambda g: pdate(g.get('fecha')) or datetime.datetime.min)

    wcc('RENDICIONES (GASTOS DEL REPORTE DE EGRESOS, periodo del EECC)', '', '', '', '', True)
    wcc('Fecha', 'Motivo', 'Entregado a', 'Monto', '', True)
    total_gastos = 0.0
    for g in gastos_periodo:
        monto_g = float(g.get('monto') or 0)
        wcc(g.get('fecha', ''), str(g.get('motivo', ''))[:80], g.get('entregado_a', ''), round(monto_g, 2), fmt_cols=(4,))
        total_gastos += monto_g
    wcc('TOTAL GASTOS', '', '', round(total_gastos, 2), '', True, catfill, fmt_cols=(4,))
    rc2 += 1

    wcc('REPOSICIONES BANCARIAS (cargos categoria CAJA CHICA, todas las cuentas soles)', '', '', '', '', True)
    wcc('Fecha', '', '', 'Monto', '', True)
    total_repuesto = 0.0
    for fecha_op, monto in sorted(caja_chica_repuesto_rows, key=lambda x: pdate(x[0]) or datetime.datetime.min):
        wcc(fecha_op, '', '', round(monto, 2), fmt_cols=(4,))
        total_repuesto += monto
    wcc('TOTAL REPUESTO', '', '', round(total_repuesto, 2), '', True, catfill, fmt_cols=(4,))
    rc2 += 1

    diferencia_gr = round(total_repuesto - total_gastos, 2)
    wcc('DIFERENCIA (REPUESTO - GASTOS)', '', '', diferencia_gr,
        'Informativo: la reposicion es un monto FIJO semanal, no un reembolso exacto de lo gastado - una diferencia no es error por si sola.',
        True, fmt_cols=(4,))

    n_semanas = _semanas_en_periodo(periodo_ini, periodo_fin)
    cadencia_esperada = round(reposicion_semanal * n_semanas, 2)
    diferencia_cadencia = round(total_repuesto - cadencia_esperada, 2)
    alerta_cadencia = reposicion_semanal > 0 and abs(diferencia_cadencia) > reposicion_semanal
    wcc('CADENCIA ESPERADA (reposicion_semanal x semanas del periodo)', '', '', cadencia_esperada,
        f'{n_semanas} semana(s) x S/{reposicion_semanal:,.2f} vs repuesto real S/{total_repuesto:,.2f} (dif S/{diferencia_cadencia:,.2f})',
        True, pendfill if alerta_cadencia else okfill, fmt_cols=(4,))
    rc2 += 1
    if alerta_cadencia:
        pendientes_caja_chica.append({
            'fecha': periodo_fin.strftime('%d/%m/%Y') if periodo_fin else None,
            'monto': diferencia_cadencia,
            'descripcion': (f'CAJA CHICA {local_nombre} ({local_resp}): repuesto S/{total_repuesto:,.2f} vs cadencia '
                             f'esperada S/{cadencia_esperada:,.2f} ({n_semanas} semana(s) x S/{reposicion_semanal:,.2f})'),
            'motivo': 'CAJA CHICA CADENCIA DE REPOSICION FUERA DE LO ESPERADO',
        })

    # ---- Depositos de venta en efectivo (reporte) vs abonos del banco ----
    depositos_reporte = sorted(EGRESOS_DATA.get('depositos') or [],
                                key=lambda d: pdate(d.get('fecha')) or datetime.datetime.min)
    sin_match_ids = {id(d) for d in DEPOSITOS_SIN_MATCH}
    total_depositos_reporte = round(sum(float(d.get('monto') or 0) for d in depositos_reporte), 2)
    n_cruzaron = len(depositos_reporte) - len(DEPOSITOS_SIN_MATCH)
    wcc('DEPOSITOS DE VENTA EN EFECTIVO (REPORTE) vs ABONOS DEL BANCO', '', '', '', '', True, tfill)
    wcc('Fecha', 'Motivo', 'Cruzo con abono', 'Monto', '', True)
    for d in depositos_reporte:
        cruzo = id(d) not in sin_match_ids
        wcc(d.get('fecha', ''), str(d.get('motivo', ''))[:60], 'SI' if cruzo else 'NO', round(float(d.get('monto') or 0), 2),
            fill=None if cruzo else pendfill, fmt_cols=(4,))
    wcc('TOTAL DEPOSITOS DEL REPORTE', '', '', total_depositos_reporte, '', True, catfill, fmt_cols=(4,))
    wcc('Cruzaron con un abono', f'{n_cruzaron}/{len(depositos_reporte)}', '', '', '', True,
        okfill if depositos_reporte and n_cruzaron == len(depositos_reporte) else pendfill)
    rc2 += 1

    if DEPOSITOS_SIN_MATCH:
        wcc('PENDIENTE: depositos del reporte SIN abono que calce (monto exacto +/-0.05, fecha +/-1 dia)', '', '', '', '', True, pendfill)
        for d in DEPOSITOS_SIN_MATCH:
            wcc(d.get('fecha', ''), str(d.get('motivo', ''))[:60], '', round(float(d.get('monto') or 0), 2), fmt_cols=(4,))
            pendientes_caja_chica.append({
                'fecha': d.get('fecha'),
                'monto': round(float(d.get('monto') or 0), 2),
                'descripcion': (f"CAJA CHICA {local_nombre}: deposito de venta en efectivo del reporte sin abono "
                                f"que calce ({d.get('fecha')} S/{float(d.get('monto') or 0):,.2f})"),
                'motivo': 'DEPOSITO CAJA CHICA SIN ABONO',
            })
        rc2 += 1

    if ABONOS_DEPOSITO_SIN_MATCH:
        wcc('PENDIENTE: abonos tipo deposito/ingreso en efectivo SIN deposito del reporte que los reclame', '', '', '', '', True, pendfill)
        for m in ABONOS_DEPOSITO_SIN_MATCH:
            wcc(m['fop'], str(m['desc'])[:60], '', round(float(m['abono']), 2), fmt_cols=(4,))
            pendientes_caja_chica.append({
                'fecha': m['fop'],
                'monto': round(float(m['abono']), 2),
                'descripcion': (f"CAJA CHICA {local_nombre}: abono {m['fop']} S/{float(m['abono']):,.2f} "
                                f"({str(m['desc'])[:40]}) sin deposito del reporte de egresos que lo reclame"),
                'motivo': 'ABONO DEPOSITO SIN REPORTE',
            })
        rc2 += 1

    sin_rendicion = False
    if periodo_fin:
        ventana_ini = periodo_fin - datetime.timedelta(days=7)
        hay_gasto_reciente = any(
            pdate(g.get('fecha')) and ventana_ini <= pdate(g.get('fecha')) <= periodo_fin for g in gastos_periodo
        )
        if not hay_gasto_reciente:
            sin_rendicion = True
            wcc('ALERTA', 'SIN RENDICION ESTA SEMANA', '', '',
                f'No hay gastos del reporte de egresos entre {ventana_ini.strftime("%d/%m/%Y")} y {periodo_fin.strftime("%d/%m/%Y")}.',
                True, pendfill)
            pendientes_caja_chica.append({
                'fecha': periodo_fin.strftime('%d/%m/%Y'),
                'monto': 0,
                'descripcion': f'CAJA CHICA {local_nombre} ({local_resp}): sin gastos en el reporte en los ultimos 7 dias del periodo',
                'motivo': 'CAJA CHICA SIN RENDIR',
            })
else:
    # -----------------------------------------------------------------------
    # Regla vieja, vigente hasta jul-2026 (sin --egresos): fondo fijo S/500 +
    # boletas rendidas del CSV de comprobantes (columna CAJA_CHICA=SI) +
    # reposiciones bancarias (cargos categoria CAJA CHICA). NO TOCAR: es el
    # comportamiento que la regresion de junio verifica bit a bit.
    # -----------------------------------------------------------------------
    wcc('Local', local_nombre, 'Responsable', local_resp, 'Aritmetica: SALDO TEORICO = FONDO FIJO - BOLETAS RENDIDAS + REPOSICIONES BANCARIAS', True)
    rc2 += 1
    wcc('FONDO FIJO (no cambia)', round(FONDO_CAJA_CHICA, 2), '', '', 'Monto fijo asignado al local, S/500 por local (decision del negocio)', True, tfill, fmt_cols=(2,))
    rc2 += 1

    boletas = []
    if COMPROBANTES_CARGADOS and local_nombre:
        for c in comprobantes:
            if not c['_CAJA_CHICA']: continue
            if norm(c.get('LOCAL', '')) != norm(local_nombre): continue
            fe = c['_FECHA_EMISION']
            if periodo_ini and periodo_fin and fe and not (periodo_ini <= fe <= periodo_fin): continue
            boletas.append(c)
    boletas.sort(key=lambda c: c['_FECHA_EMISION'] or datetime.datetime.min)

    wcc('BOLETAS RENDIDAS (periodo del EECC)', '', '', '', '', True)
    wcc('Fecha', 'Proveedor', 'Serie', 'Monto', '', True)
    total_rendido = 0
    for c in boletas:
        fe_str = c.get('FECHA_EMISION', '')
        wcc(fe_str, c.get('PROVEEDOR', ''), c.get('SERIE_NUMERO', ''), round(c['_TOTAL'], 2), fmt_cols=(4,))
        total_rendido += c['_TOTAL']
    wcc('TOTAL RENDIDO', '', '', round(total_rendido, 2), '', True, catfill, fmt_cols=(4,))
    rc2 += 1

    wcc('REPOSICIONES BANCARIAS (cargos categoria CAJA CHICA, todas las cuentas soles)', '', '', '', '', True)
    wcc('Fecha', '', '', 'Monto', '', True)
    total_repuesto = 0
    for fecha_op, monto in sorted(caja_chica_repuesto_rows, key=lambda x: pdate(x[0]) or datetime.datetime.min):
        wcc(fecha_op, '', '', round(monto, 2), fmt_cols=(4,))
        total_repuesto += monto
    wcc('TOTAL REPUESTO', '', '', round(total_repuesto, 2), '', True, catfill, fmt_cols=(4,))
    rc2 += 1

    saldo_teorico = round(FONDO_CAJA_CHICA - total_rendido + total_repuesto, 2)
    diferencia = round(saldo_teorico - FONDO_CAJA_CHICA, 2)
    wcc('SALDO TEORICO', '', '', saldo_teorico, 'FONDO - RENDIDO + REPUESTO. Debe volver a S/500 si la reposicion cubrio exactamente lo rendido.',
        True, okfill if abs(diferencia) < 0.01 else pendfill, fmt_cols=(4,))
    wcc('DIFERENCIA vs FONDO (S/500)', '', '', diferencia,
        'Negativo = falta reponer a la caja; positivo = se repuso de mas o hay boletas de un periodo anterior sin registrar.',
        True, fmt_cols=(4,))
    rc2 += 1

    sin_rendicion = False
    if periodo_fin:
        ventana_ini = periodo_fin - datetime.timedelta(days=7)
        hay_boleta_reciente = any(c['_FECHA_EMISION'] and ventana_ini <= c['_FECHA_EMISION'] <= periodo_fin for c in boletas)
        if not hay_boleta_reciente:
            sin_rendicion = True
            wcc('ALERTA', 'SIN RENDICION ESTA SEMANA', '', '',
                f'No hay boletas de caja chica rendidas entre {ventana_ini.strftime("%d/%m/%Y")} y {periodo_fin.strftime("%d/%m/%Y")}.',
                True, pendfill)
            pendientes_caja_chica.append({
                'fecha': periodo_fin.strftime('%d/%m/%Y'),
                'monto': 0,
                'descripcion': f'CAJA CHICA {local_nombre} ({local_resp}): sin rendicion en los ultimos 7 dias del periodo',
                'motivo': 'CAJA CHICA SIN RENDIR',
            })
    if not COMPROBANTES_CARGADOS:
        wcc('NOTA', 'No se cargo --comprobantes en esta corrida: BOLETAS RENDIDAS queda en 0 por falta de datos, no por falta real de rendicion.', '', '', '', False)

for rr in range(1, rc2):
    for cc3 in range(1, 6): scc.cell(rr, cc3).border = border

# ---------------------------------------------------------------------------
# VERIFICACION
# ---------------------------------------------------------------------------
sv = wb.create_sheet('VERIFICACION')
sv.column_dimensions['A'].width = 42; sv.column_dimensions['B'].width = 18; sv.column_dimensions['C'].width = 56
rv = 1


def wv(a, b='', note='', bold=False, fill=None, fmt=False):
    global rv
    ca = sv.cell(rv, 1, a); cb = sv.cell(rv, 2, b); cc4 = sv.cell(rv, 3, note)
    ca.font = Font(bold=bold, size=10); cb.font = Font(bold=bold, size=10); cc4.font = Font(size=8, italic=True)
    if fmt and b != '': cb.number_format = '#,##0.00'
    if fill: ca.fill = cb.fill = fill
    rv += 1


wv(f'VERIFICACION DE CUADRE - {EMP} - {PERIODO_LABEL}', '', '', True, tfill); rv += 1

if HEREDAR_MAP is not None:
    wv('HEREDAR CATEGORIAS DE CORRIDA ANTERIOR', os.path.basename(args.heredar),
       'ver scripts/heredar_categorias.py - Proveedor/Categoria heredados cuando esta corrida no los determino sola', True, tfill)
    wv('Cargos que heredaron Proveedor', HEREDAR_STATS['prov'])
    wv('Cargos que heredaron Categoria', HEREDAR_STATS['cat'])
    if HEREDAR_STATS['reclasificaciones']:
        wv('Cargos reclasificados vs archivo anterior', len(HEREDAR_STATS['reclasificaciones']),
           'la corrida nueva SI determino categoria propia y difiere de la del archivo anterior (ver consola/reporte de la corrida)', True, pendfill)
    rv += 1

# ---- Cuadre EXACTO por cuenta (nuevo v3.1): SALDO INICIAL + ABONOS - CARGOS = SALDO FINAL ----
wv('CUADRE EXACTO POR CUENTA (v3.1)', '', 'saldo inicial + abonos - cargos = saldo final, segun el propio EECC', True, tfill)
for r in resultados:
    c = r['cuenta']; m = c['meta']; alias = c['alias'] or 'PRINCIPAL'
    nombre_cuenta = f"Cuenta {alias} ({m['moneda']}, {m['banco']}, ...{(m.get('cuenta') or '')[-4:]})"
    if r.get('vacia'):
        wv(nombre_cuenta, 'SIN MOVIMIENTOS', f"EECC {os.path.basename(c['path'])} - sin movimientos en el periodo (todo en 0)", True, pendfill)
        continue
    rc = r['cargos']; ra = r['abonos']
    calc_final = round((m['saldo_inicial'] or 0) + ra['total'] - rc['total'], 2)
    reportado_final = m['saldo_final']
    ok = m['anchor_ok'] and reportado_final is not None and abs(calc_final - reportado_final) < 0.01
    wv(nombre_cuenta, 'OK' if ok else 'ERROR',
       f"ini {m['saldo_inicial']:,.2f} + abonos {ra['total']:,.2f} - cargos {rc['total']:,.2f} = {calc_final:,.2f} vs saldo final EECC {reportado_final:,.2f} | "
       f"{'ancla exacto' if m['anchor_ok'] else 'NO ancla (' + str(len(m['anchor_mismatches'])) + ' fila(s) con diferencia)'} | {os.path.basename(c['path'])}",
       True, okfill if ok else badfill)
rv += 1

wv('Movimientos de CARGO (cuenta(s) SOLES consolidadas)', n_cargos)
wv('Movimientos de ABONO (cuenta(s) SOLES consolidadas)', n_abonos)
rv += 1
wv('TOTAL CARGOS (col. Cargo)', round(total, 2), 'suma exacta del estado' + (' - solo SOLES' if MULTI_CUENTA else ''), True, fmt=True)
wv('TOTAL ABONOS (col. Abono)', round(totA, 2), 'suma exacta del estado' + (' - solo SOLES' if MULTI_CUENTA else ''), True, fmt=True)
neto = round(totA - total, 2)
wv('NETO DEL MES (abonos - cargos)', neto, '', True, PatternFill('solid', fgColor='E2EFDA'), fmt=True)
rv += 1
wv('Identificado (beneficiario o categoria x desc)', round(principal_res['cargos']['benef_amt'], 2),
   f"{principal_res['cargos']['benef_amt'] / principal_res['cargos']['total'] * 100:.1f}% del gasto de la cuenta principal" if principal_res['cargos']['total'] else '',
   True, PatternFill('solid', fgColor='C6EFCE'), fmt=True)
sinbenef_amt = principal_res['cargos']['sinbenef_amt']
if sinbenef_amt > 0.005:
    wv('I-BANC sin constancia que cruce (cuenta principal)', round(sinbenef_amt, 2),
       f"{sinbenef_amt / principal_res['cargos']['total'] * 100:.1f}% - buscar en Gmail (rango de fechas mas amplio)", True, pendfill, fmt=True)
else:
    wv('I-BANC sin constancia que cruce (cuenta principal)', 0, 'Gmail ya cruzado: 0 cargos sin beneficiario', True, PatternFill('solid', fgColor='C6EFCE'), fmt=True)
rv += 1
wv('Con categoria asignada (SOLES)', round(explic, 2), f'{explic / total * 100:.1f}% del gasto' if total else '', True, fmt=True)
wv('Sin categoria (falta decision del dueno)', round(total - explic, 2),
   f'{(total - explic) / total * 100:.1f}% - en su mayoria SILVA (en blanco por diseno)' if total else '', True, catfill, fmt=True)
rv += 1

if MULTI_CUENTA:
    wv('TRANSFERENCIAS ENTRE CUENTAS DETECTADAS (v3.1)', len(transferencias), '', True, tfill)
    for t in transferencias:
        wv(f"   S/{t['monto']:,.2f}: {t['cuenta_cargo']} ({t['fecha_cargo']}) -> {t['cuenta_abono']} ({t['fecha_abono']})",
            '', f"{t['desc_cargo'][:40]} / {t['desc_abono'][:40]}")
    rv += 1

n_cargos_data_total = n_cargos_data
pct_conc = (n_conc_si / n_cargos_data_total * 100) if n_cargos_data_total else 0
wv('CONCILIACION DE COMPROBANTES (v3) - todas las cuentas', '', '', True, tfill)
wv('Cargos CONCILIADO = SI', n_conc_si, f'{n_conc_si}/{n_cargos_data_total} = {pct_conc:.1f}% (meta 100%)', True,
   okfill if pct_conc >= 99.99 else pendfill)
wv('Cargos CONCILIADO = NO', n_cargos_data_total - n_conc_si)
if not COMPROBANTES_CARGADOS:
    wv('Nota', '', 'Corrida SIN --comprobantes: solo se resolvieron ITF/COMISION, traspasos entre hermanas/cuentas y caja chica. Cargar el CSV de facturas para conciliar el resto.', False)
for motivo_key in sorted(motivo_counts):
    wv('   Motivo: ' + motivo_key, motivo_counts[motivo_key])
for cc5 in range(1, 4):
    for rr in range(1, rv): sv.cell(rr, cc5).border = border

# ---------------------------------------------------------------------------
# Orden de hojas, mayusculas, hoja activa
# ---------------------------------------------------------------------------
order = ['CARGOS'] + [f'CARGOS {a}' for a in sheet_aliases if a] + \
        ['ABONOS'] + [f'ABONOS {a}' for a in sheet_aliases if a] + \
        ['FLUJO CAJA', 'EGP', 'CAJA CHICA', 'VERIFICACION', 'EECC'] + \
        [f'EECC {a}' for a in sheet_aliases if a]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
for _ws in wb.worksheets: _ws.sheet_view.tabSelected = False
wb['CARGOS'].sheet_view.tabSelected = True; wb.active = 0
for sh in wb.worksheets:
    if sh.title.startswith('EECC'): continue
    for rr in sh.iter_rows():
        for cell in rr:
            if isinstance(cell.value, str): cell.value = cell.value.upper()
wb.save(dst)

if HEREDAR_MAP is not None:
    print(f"HEREDAR ({args.heredar}): Proveedor heredado en {HEREDAR_STATS['prov']} cargo(s), "
          f"Categoria heredada en {HEREDAR_STATS['cat']} cargo(s).")
    if HEREDAR_STATS['reclasificaciones']:
        print(f"Reclasificaciones vs archivo anterior ({len(HEREDAR_STATS['reclasificaciones'])}):")
        for rcl in HEREDAR_STATS['reclasificaciones']:
            print(f"  {rcl['sheet']} {rcl['fecha']} S/{rcl['monto']:.2f} {rcl['proveedor']}: "
                  f"{rcl['categoria_anterior']} -> {rcl['categoria_nueva']}")

# ---------------------------------------------------------------------------
# pendientes.json (para notificar_pendientes.py)
# ---------------------------------------------------------------------------
if args.pendientes:
    pend_compras = []
    pend_conciliacion = list(pendientes_caja_chica)
    # Releer TODAS las hojas CARGOS/CARGOS <ALIAS> recien escritas (ya en
    # mayusculas) para construir el detalle sin duplicar logica.
    for r in resultados:
        if r.get('vacia'):
            continue
        sc = r['cargos']['sheet']; sc_last = r['cargos']['last_row']
        for rr in range(2, sc_last + 1):
            conc = sc.cell(rr, 17).value
            if conc == 'SI': continue
            motivo_v = sc.cell(rr, 18).value or ''
            item = {
                'fecha': sc.cell(rr, 1).value,
                'monto': sc.cell(rr, 9).value,
                'descripcion': f"{sc.cell(rr, 13).value or ''} - {sc.cell(rr, 12).value or ''} ({sc.cell(rr, 7).value or ''})",
                'motivo': motivo_v,
            }
            if motivo_v == 'SIN COMPROBANTE':
                pend_compras.append(item)
            elif motivo_v in ('SIN CONSTANCIA', 'PENDIENTE CONSTANCIA'):
                pend_conciliacion.append(item)

    pend_pagos = []
    if COMPROBANTES_CARGADOS:
        pend_por_prov = {}
        for c in comprobantes:
            if norm(c.get('ESTADO_PAGO', '')) != 'PENDIENTE':
                continue
            key = c.get('PROVEEDOR', '') or 'SIN PROVEEDOR'
            pend_por_prov.setdefault(key, {'monto': 0, 'n': 0})
            pend_por_prov[key]['monto'] += c['_TOTAL'] or 0
            pend_por_prov[key]['n'] += 1
        for prov_name, info in sorted(pend_por_prov.items()):
            pend_pagos.append({
                'fecha': None,
                'monto': round(info['monto'], 2),
                'descripcion': f"{prov_name}: {info['n']} factura(s) pendiente(s) de pago",
                'motivo': 'FACTURAS PENDIENTES DE PAGO',
            })

    pendientes = {
        'compras': pend_compras,
        'pagos': pend_pagos,
        'conciliacion': pend_conciliacion,
        'ventas': [],  # pendiente: requiere reportes de ventas/POS (Fase 4.3, sin ejemplos aun)
    }
    with open(args.pendientes, 'w', encoding='utf-8') as f:
        json.dump(pendientes, f, ensure_ascii=False, indent=2)

print(f'== {EMP} ({PERIODO_LABEL}) ==')
print(f'Guardado: {dst}')
for r in resultados:
    c = r['cuenta']; m = c['meta']; alias = c['alias'] or 'PRINCIPAL'
    if r.get('vacia'):
        print(f"  Cuenta {alias}: SIN MOVIMIENTOS ({os.path.basename(c['path'])})")
        continue
    rc = r['cargos']; ra = r['abonos']
    print(f"  Cuenta {alias} ({m['moneda']}, {m['banco']}): {rc['n_cargos_data']} cargos S/{rc['total']:,.2f} + "
          f"{ra['n_abonos']} abonos S/{ra['total']:,.2f} | saldo ini {m['saldo_inicial']} -> fin {m['saldo_final']} "
          f"| ancla: {'OK' if m['anchor_ok'] else 'ERROR'}")
if transferencias:
    print(f'TRANSFERENCIAS ENTRE CUENTAS detectadas: {len(transferencias)}')
    for t in transferencias:
        print(f"  S/{t['monto']:,.2f}  {t['cuenta_cargo']} ({t['fecha_cargo']}) -> {t['cuenta_abono']} ({t['fecha_abono']})")
print(f'Cargos (SOLES consolidado): S/{total:,.2f} | Abonos: S/{totA:,.2f} | Neto: S/{neto:,.2f}')
print(f'EXPLICADO: S/{explic:,.2f} de S/{total:,.2f} ({explic / total * 100:.1f}%)' if total else 'EXPLICADO: N/A')
print(f'CONCILIADO (comprobantes, todas las cuentas): {n_conc_si}/{n_cargos_data_total} = {pct_conc:.1f}%')
for mk in sorted(motivo_counts): print(f'  MOTIVO {mk}: {motivo_counts[mk]}')
pend_amt = sum(v for *_, v in pend_display)
print(f'PENDIENTE cuenta principal (sin categoria): {len(pend_display)} movs = S/{pend_amt:,.2f}')
for f, m, dd, v in sorted(pend_display, key=lambda x: -x[3])[:60]:
    print(f'  {f} S/{v:>9,.2f} {N(m)[:20]:20} {dd}')
