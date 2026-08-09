#!/usr/bin/env python3
"""Parsers de estados de cuenta (EECC) - SCONCHA conciliacion v3.1.

Fuente de verdad del skill (ver ../SKILL.md). Modulo compartido usado por
build_conciliacion.py para aceptar los formatos OFICIALES del banco (ademas del
export Excel "Movimientos_..." que v3 ya soportaba):

- parse_eecc_interbank_xlsx: export Excel "Movimientos_[cuenta]_..." (hoja 'Page 1',
  encabezado fila 12, datos desde fila 13). Es el formato v3 original.
- parse_eecc_interbank_pdf: PDF oficial del estado de cuenta Interbank (texto
  extraible con pypdf). NUEVO en v3.1.
- parse_eecc_bbva_html: archivo .xls de BBVA que en realidad es HTML (BOM utf-8-sig
  + <html>). NUEVO en v3.1.

Salida normalizada comun para los tres parsers:
  movimientos: lista de dicts con las claves que ya consume build_conciliacion.py:
    {'fop': 'DD/MM/AAAA', 'fpr': 'DD/MM/AAAA', 'nro': str|None, 'mov': str,
     'desc': str, 'canal': str, 'cargo': float|None, 'abono': float|None,
     'saldo': float, '_id': int}
    ('mov' es el codigo corto de "Movimiento" de la hoja xlsx; para PDF/BBVA no
    existe ese campo separado, se deja '' y todo va en 'desc' - cat_by_desc() en
    build_conciliacion.py ya concatena mov+desc, asi que funciona igual.)
  meta: dict con:
    {'banco': 'INTERBANK'|'BBVA', 'formato': 'xlsx'|'pdf'|'bbva_html',
     'cuenta': str, 'moneda': 'PEN'|'USD', 'cliente': str,
     'periodo_mes': int|None, 'periodo_anio': int|None,
     'saldo_inicial': float|None, 'saldo_final': float|None,
     'anchor_ok': bool, 'anchor_mismatches': [ {..} ],
     'n_movimientos': int, 'unparsed_lines': [str, ...]}

Cada parser VALIDA la aritmetica del saldo corrido (saldo_anterior + abono - cargo
== saldo_actual, fila a fila, en el orden real del libro mayor -no el de fecha-)
y reporta si "ancla" exacto: los PDF de Interbank y el HTML de BBVA SIEMPRE deben
anclar exacto (son el saldo contable real del banco); el xlsx exportado mantiene
su comportamiento de v2/v3 (no siempre ancla, porque los abonos IZIPAY/POS se
listan fuera de orden de saldo dentro del export -ver anchor_ok=False y el detalle
en anchor_mismatches).
"""
import re
import os
import datetime
import html.parser as _htmlparser

import openpyxl

MESES_ES = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO',
            'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
MES_NUM = {m: i + 1 for i, m in enumerate(MESES_ES)}

_id_counter = [0]


def _next_id():
    _id_counter[0] += 1
    return _id_counter[0]


def _fnum(s):
    """'-14,953.97' / '1,500.00' / '' -> float o None."""
    s = str(s or '').strip()
    if not s:
        return None
    try:
        return float(s.replace(',', ''))
    except Exception:
        return None


def guess_period_from_filename(path):
    """Ultimo recurso: los archivos siguen el patron ..._MMAAAA.ext (ej.
    EC_4134_062026.pdf, EC_BBVA_8579_062026.xls) -> (mes, anio)."""
    m = re.search(r'_(\d{2})(\d{4})\.\w+$', os.path.basename(path))
    if m:
        mes, anio = int(m.group(1)), int(m.group(2))
        if 1 <= mes <= 12:
            return mes, anio
    return None, None


def detect_format(path):
    """Detecta el formato por extension + contenido (el .xls de BBVA es HTML
    disfrazado, hay que mirar los primeros bytes, no confiar en la extension)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.pdf':
        return 'pdf'
    if ext in ('.xlsx', '.xlsm'):
        return 'xlsx'
    if ext == '.xls':
        with open(path, 'rb') as f:
            head = f.read(200)
        head_txt = head.lstrip(b'\xef\xbb\xbf').lstrip()
        if head_txt.lower().startswith(b'<html') or b'<html' in head[:100].lower():
            return 'bbva_html'
        # .xls real (BIFF) - no soportado (xlrd no sirve para openpyxl); lo
        # dejamos pasar como error explicito en parse_eecc().
        return 'xls_biff_unsupported'
    raise ValueError(f'Extension no reconocida para EECC: {path}')


# ---------------------------------------------------------------------------
# 1) Interbank - export Excel "Movimientos_..." (formato v3 original)
# ---------------------------------------------------------------------------
def parse_eecc_interbank_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Page 1'] if 'Page 1' in wb.sheetnames else wb.active

    movimientos = []
    for r in range(13, ws.max_row + 1):
        fop = ws.cell(r, 2).value
        mov = ws.cell(r, 5).value
        cargo = ws.cell(r, 8).value
        abono = ws.cell(r, 9).value
        if fop is None and mov is None and cargo is None and abono is None:
            continue
        movimientos.append({
            'fop': fop, 'fpr': ws.cell(r, 3).value, 'nro': ws.cell(r, 4).value,
            'mov': mov, 'desc': ws.cell(r, 6).value,
            'canal': (str(ws.cell(r, 7).value).strip() if ws.cell(r, 7).value else ''),
            'cargo': (abs(float(cargo)) if cargo not in (None, '') else None),
            'abono': (abs(float(abono)) if abono not in (None, '') else None),
            'saldo': (float(ws.cell(r, 10).value) if ws.cell(r, 10).value not in (None, '') else None),
            '_id': _next_id(),
        })

    # Reconstruccion de cadena de saldo (igual que v3: el export IZIPAY/POS no
    # siempre viene ordenado por saldo -> se reconstruye con Counter, sin asumir
    # orden de filas).
    from collections import Counter
    after = Counter()
    before = Counter()
    for d in movimientos:
        e = round((d['abono'] or 0) - (d['cargo'] or 0), 2)
        s = round(d['saldo'] or 0, 2)
        after[s] += 1
        before[round(s - e, 2)] += 1
    op = list((before - after).elements())
    cl = list((after - before).elements())
    anchor_ok = (len(op) == 1 and len(cl) == 1)
    saldo_inicial = round(op[0], 2) if anchor_ok else None
    saldo_final = round(cl[0], 2) if anchor_ok else None

    mes, anio = None, None
    _dates = [d['fop'] for d in movimientos if d['fop']]
    if _dates:
        try:
            parsed = [d if isinstance(d, datetime.datetime) else
                      datetime.datetime.strptime(str(d), '%d/%m/%Y') for d in _dates]
            cnt = Counter((d.month, d.year) for d in parsed)
            mes, anio = cnt.most_common(1)[0][0]
        except Exception:
            mes, anio = guess_period_from_filename(path)
    else:
        mes, anio = guess_period_from_filename(path)

    # Normalizar fop/fpr a texto 'DD/MM/AAAA' (a veces openpyxl los entrega como
    # datetime si la celda tiene formato fecha).
    for d in movimientos:
        for k in ('fop', 'fpr'):
            v = d[k]
            if isinstance(v, datetime.datetime):
                d[k] = v.strftime('%d/%m/%Y')

    meta = {
        'banco': 'INTERBANK', 'formato': 'xlsx', 'cuenta': '', 'moneda': 'PEN',
        'cliente': '', 'periodo_mes': mes, 'periodo_anio': anio,
        'saldo_inicial': saldo_inicial, 'saldo_final': saldo_final,
        'anchor_ok': anchor_ok, 'anchor_mismatches': [],
        'n_movimientos': len(movimientos), 'unparsed_lines': [],
    }
    return movimientos, meta


# ---------------------------------------------------------------------------
# 2) Interbank - PDF oficial del estado de cuenta
# ---------------------------------------------------------------------------
CANALES_CONOCIDOS = {'INTERNO', 'WEB', 'IZIPAY', 'I-BANC', 'DEVOLUCION'}
CANALES_PREFIJO = {'TIENDA', 'ATM', 'AGENTE'}  # van seguidos de un codigo: "TIENDA 164"

_MOV_LINE_RE = re.compile(
    r'^(\d{2}/\d{2})\s+(\d{2}/\d{2})\s+(.*?)\s+(-?[\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$'
)


def _split_canal(rest):
    """Separa canal de la descripcion en una linea de movimiento Interbank PDF.
    Ver docstring del modulo / SKILL.md para el detalle de canales vistos."""
    tokens = rest.split()
    if not tokens:
        return '', rest
    if tokens[-1] in CANALES_CONOCIDOS:
        return tokens[-1], ' '.join(tokens[:-1])
    if len(tokens) >= 2 and tokens[-2] in CANALES_PREFIJO:
        return f'{tokens[-2]} {tokens[-1]}', ' '.join(tokens[:-2])
    canal = ''
    if re.search(r'\bCCE\b', rest):
        canal = 'CCE'
    elif re.search(r'\bPOS\b', rest):
        canal = 'POS'
    return canal, rest


def _extract_nro(desc):
    m = re.search(r'(\d{5,})\s*$', desc)
    return m.group(1) if m else None


def _dm_to_ddmmyyyy(dm, mes_stmt, anio_stmt):
    dd, mm = dm.split('/')
    mm_i = int(mm)
    anio = anio_stmt
    if mes_stmt and mm_i > mes_stmt and (mm_i - mes_stmt) >= 6:
        anio -= 1  # ej. operacion de diciembre en un estado de cuenta de enero
    return f'{dd}/{mm}/{anio}' if anio else f'{dd}/{mm}'


def parse_eecc_interbank_pdf(path, password=None):
    """password: contrasena del PDF si Interbank lo protegio (jul-2026: los
    PDF de "Cuenta Negocio" empezaron a venir cifrados con el RUC del titular
    como contrasena; los de jun-2026 no la traian, asi que este caso no
    se habia visto hasta ahora). Si el PDF no esta cifrado, el parametro se
    ignora sin error."""
    import pypdf
    reader = pypdf.PdfReader(path)
    if reader.is_encrypted:
        if not password:
            raise ValueError(
                f'{path}: el PDF esta protegido con contrasena y no se paso ninguna '
                f'(--pdf-password). Interbank suele usar el RUC del titular de la '
                f'cuenta como contrasena de los EECC de "Cuenta Negocio".'
            )
        resultado = reader.decrypt(password)
        if resultado == 0:  # 0 = ni user ni owner password calzaron
            raise ValueError(
                f'{path}: la contrasena provista no abrio el PDF cifrado '
                f'(se intento como --pdf-password).'
            )
    full_text = '\n'.join((p.extract_text() or '') for p in reader.pages)

    m = re.search(r'Mes:\s*([A-Za-z\xc0-\xff]+)\s+(\d{4})', full_text)
    mes = anio = None
    if m:
        mes = MES_NUM.get(m.group(1).strip().upper())
        anio = int(m.group(2))
    if not mes:
        mes, anio = guess_period_from_filename(path)

    cuenta = ''
    mc = re.search(r'CUENTA:\s*([\d\-]+)', full_text)
    if mc:
        cuenta = mc.group(1).strip()
    cliente = ''
    mcli = re.search(r'CLIENTE:\s*(.+)', full_text)
    if mcli:
        cliente = mcli.group(1).strip()
    moneda = 'PEN'
    mm2 = re.search(r'MONEDA:\s*(\S+)', full_text)
    if mm2 and 'SOL' not in mm2.group(1).upper():
        moneda = 'USD'

    movimientos = []
    unparsed = []
    for line in full_text.splitlines():
        line = line.strip()
        if not line:
            continue
        mline = _MOV_LINE_RE.match(line)
        if not mline:
            continue
        d1, d2, rest, monto_s, saldo_s = mline.groups()
        monto = _fnum(monto_s)
        saldo = _fnum(saldo_s)
        if monto is None or saldo is None:
            unparsed.append(line)
            continue
        canal, desc = _split_canal(rest)
        nro = _extract_nro(desc)
        movimientos.append({
            'fop': _dm_to_ddmmyyyy(d1, mes, anio),
            'fpr': _dm_to_ddmmyyyy(d2, mes, anio),
            'nro': nro, 'mov': '', 'desc': desc, 'canal': canal,
            'cargo': abs(monto) if monto < 0 else None,
            'abono': monto if monto >= 0 else None,
            'saldo': saldo,
            '_id': _next_id(),
        })

    saldo_inicial = saldo_final = None
    anchor_ok = True
    mismatches = []
    if movimientos:
        first = movimientos[0]
        saldo_inicial = round(first['saldo'] - (first['abono'] or 0) + (first['cargo'] or 0), 2)
        prev = saldo_inicial
        for d in movimientos:
            expected = round(prev + (d['abono'] or 0) - (d['cargo'] or 0), 2)
            if abs(expected - round(d['saldo'], 2)) > 0.01:
                anchor_ok = False
                mismatches.append({
                    'fecha': d['fop'], 'desc': d['desc'], 'esperado': expected, 'reportado': d['saldo'],
                })
            prev = d['saldo']
        saldo_final = movimientos[-1]['saldo']

    meta = {
        'banco': 'INTERBANK', 'formato': 'pdf', 'cuenta': cuenta, 'moneda': moneda,
        'cliente': cliente, 'periodo_mes': mes, 'periodo_anio': anio,
        'saldo_inicial': saldo_inicial, 'saldo_final': saldo_final,
        'anchor_ok': anchor_ok, 'anchor_mismatches': mismatches,
        'n_movimientos': len(movimientos), 'unparsed_lines': unparsed,
    }
    return movimientos, meta


# ---------------------------------------------------------------------------
# 3) BBVA - archivo .xls que en realidad es HTML (BOM utf-8-sig + <html>)
# ---------------------------------------------------------------------------
class _TableRowParser(_htmlparser.HTMLParser):
    """Extrae todas las filas <tr> con sus <td> como texto plano, de cualquier
    tabla del documento (no distingue tablas anidadas; luego filtramos por
    numero de columnas =9, que es la tabla de movimientos)."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.rows = []
        self.cur = None
        self.cell = None

    def handle_starttag(self, tag, attrs):
        if tag == 'tr':
            self.cur = []
        elif tag == 'td':
            self.cell = ''

    def handle_data(self, data):
        if self.cell is not None:
            self.cell += data

    def handle_endtag(self, tag):
        if tag == 'td' and self.cell is not None:
            if self.cur is not None:
                self.cur.append(re.sub(r'[\s\xa0]+', ' ', self.cell).strip())
            self.cell = None
        elif tag == 'tr':
            if self.cur:
                self.rows.append(self.cur)
            self.cur = None


def _dashdate_to_ddmmyyyy(dm, mes_stmt, anio_stmt):
    dd, mm = dm.split('-')
    return _dm_to_ddmmyyyy(f'{dd}/{mm}', mes_stmt, anio_stmt)


def parse_eecc_bbva_html(path):
    import html as _html
    with open(path, encoding='utf-8-sig', errors='replace') as f:
        html_text = f.read()

    mes, anio = guess_period_from_filename(path)

    # BBVA no trae una etiqueta "CUENTA:" explicita en este export (a diferencia
    # de Interbank); el numero de cuenta completo aparece dentro de la
    # descripcion "DEPOS. EN CTA. 00110353010003XXXX" -> usamos ese, y si no
    # aparece, el nombre de archivo (EC_BBVA_XXXX_MMAAAA.xls) trae los ultimos
    # 4 digitos como fallback.
    cuenta = ''
    mc2 = re.search(r'DEPOS\.\s*EN\s*CTA\.\s*(\d{10,})', html_text, re.I)
    if mc2:
        cuenta = mc2.group(1)
    else:
        mfn = re.search(r'_BBVA_(\d+)_', os.path.basename(path), re.I)
        cuenta = mfn.group(1) if mfn else ''
    empresa = ''
    me = re.search(r'EMPRESA:</strong>&nbsp;&nbsp;([^<\n]+)', html_text)
    if me:
        empresa = _html.unescape(me.group(1)).strip()
    moneda = 'PEN'
    mmon = re.search(r'MONEDA:\s*</b>\s*([A-Za-z\xc0-\xff]+)|MONEDA:\s*([A-Za-z\xc0-\xff]+)', html_text)
    if mmon:
        val = (mmon.group(1) or mmon.group(2) or '').upper()
        if 'SOL' not in val:
            moneda = 'USD'

    parser = _TableRowParser()
    parser.feed(html_text)
    rows = [r for r in parser.rows if len(r) == 9]

    movimientos = []
    unparsed = []
    saw_saldo_anterior = False
    stop = False
    for i, row in enumerate(rows):
        if i == 0:
            continue  # encabezado de columnas
        fop, fval, desc, oficina, canal, nro, cargo_abono, itf_s, saldo_s = row
        if not saw_saldo_anterior and desc.strip().upper() == 'SALDO ANTERIOR':
            saw_saldo_anterior = True
            continue
        if desc.strip().upper().startswith('TOTALES POR ITF'):
            stop = True
            break
        blank_amounts = not (fop.strip() or fval.strip() or canal.strip() or
                              nro.strip() or cargo_abono.strip() or saldo_s.strip())
        if blank_amounts:
            if desc.strip():
                if movimientos:
                    movimientos[-1]['desc'] += ' | ' + desc.strip()
                else:
                    unparsed.append(' | '.join(row))
            continue
        cargo_abono_v = _fnum(cargo_abono)
        itf_v = _fnum(itf_s) or 0.0
        saldo_v = _fnum(saldo_s)
        if cargo_abono_v is None or saldo_v is None or not fop.strip():
            unparsed.append(' | '.join(row))
            continue
        fop_full = _dashdate_to_ddmmyyyy(fop.strip(), mes, anio)
        fval_full = _dashdate_to_ddmmyyyy(fval.strip(), mes, anio) if fval.strip() else fop_full
        canal_v = canal.strip() or oficina.strip()
        base = {
            'fop': fop_full, 'fpr': fval_full, 'nro': nro.strip() or None,
            'mov': '', 'canal': canal_v, '_id': _next_id(),
        }
        if abs(itf_v) < 0.005:
            movimientos.append(dict(base, desc=desc.strip(),
                                     cargo=abs(cargo_abono_v) if cargo_abono_v < 0 else None,
                                     abono=cargo_abono_v if cargo_abono_v >= 0 else None,
                                     saldo=saldo_v))
        else:
            # El monto de la tabla (CARGO/ABONO) y el ITF se descuentan juntos
            # para llegar al SALDO CONTABLE reportado (verificado en los 5 EECC
            # reales de prueba); se abren en dos movimientos para que el ITF
            # quede visible como su propio cargo bancario (igual que Interbank).
            # El saldo intermedio (despues del cargo/abono, antes del ITF) no
            # viene en la tabla -> se completa en la segunda pasada de abajo.
            movimientos.append(dict(base, desc=desc.strip(),
                                     cargo=abs(cargo_abono_v) if cargo_abono_v < 0 else None,
                                     abono=cargo_abono_v if cargo_abono_v >= 0 else None,
                                     saldo=None))  # se completa abajo
            movimientos.append({
                'fop': fop_full, 'fpr': fval_full, 'nro': None, 'mov': '',
                'desc': 'ITF', 'canal': '', 'cargo': round(itf_v, 2), 'abono': None,
                'saldo': saldo_v, '_id': _next_id(),
            })

    # Segunda pasada: reconstruir saldo_inicial y completar los saldos
    # intermedios (movimientos con saldo=None, el que precede a un ITF partido).
    saldo_inicial = saldo_final = None
    anchor_ok = True
    mismatches = []
    if movimientos:
        first_known = next((d for d in movimientos if d['saldo'] is not None), None)
        if first_known is None:
            anchor_ok = False
        else:
            # reconstruir hacia atras desde el primer saldo conocido
            idx0 = movimientos.index(first_known)
            # el primer movimiento con saldo=None solo puede preceder a uno con saldo conocido
            # (siempre viene en pares consecutivos: principal (saldo=None) + ITF (saldo=conocido))
            running_known_before_first = round(
                first_known['saldo'] - (first_known['abono'] or 0) + (first_known['cargo'] or 0), 2)
            if idx0 > 0 and movimientos[idx0 - 1]['saldo'] is None:
                prev0 = movimientos[idx0 - 1]
                saldo_inicial = round(running_known_before_first - (prev0['abono'] or 0) + (prev0['cargo'] or 0), 2)
            else:
                saldo_inicial = running_known_before_first
            prev = saldo_inicial
            for d in movimientos:
                if d['saldo'] is None:
                    d['saldo'] = round(prev + (d['abono'] or 0) - (d['cargo'] or 0), 2)
                    prev = d['saldo']
                    continue
                expected = round(prev + (d['abono'] or 0) - (d['cargo'] or 0), 2)
                if abs(expected - round(d['saldo'], 2)) > 0.01:
                    anchor_ok = False
                    mismatches.append({'fecha': d['fop'], 'desc': d['desc'],
                                        'esperado': expected, 'reportado': d['saldo']})
                prev = d['saldo']
            saldo_final = movimientos[-1]['saldo']

    meta = {
        'banco': 'BBVA', 'formato': 'bbva_html', 'cuenta': cuenta, 'moneda': moneda,
        'cliente': empresa, 'periodo_mes': mes, 'periodo_anio': anio,
        'saldo_inicial': saldo_inicial, 'saldo_final': saldo_final,
        'anchor_ok': anchor_ok, 'anchor_mismatches': mismatches,
        'n_movimientos': len(movimientos), 'unparsed_lines': unparsed,
    }
    return movimientos, meta


# ---------------------------------------------------------------------------
# 4) BBVA - PDF "MOVIMIENTO Y SALDO A LA FECHA"
#
# Desde jul-2026 el estado de cuenta de BBVA llega en PDF; hasta jun-2026
# llegaba como .xls (HTML disfrazado, ver parse_eecc_bbva_html). Sin este
# parser, un PDF de BBVA caia en parse_eecc_interbank_pdf y salia con 0
# movimientos, banco 'INTERBANK' y anchor_ok=True: la conciliacion reportaba
# que todo cuadraba mientras ignoraba el mes entero de esa cuenta.
# ---------------------------------------------------------------------------
_RE_BBVA_FILA = re.compile(r'^(\d{2}-\d{2})\s+(\d{2}-\d{2})\s+(.*)$')
# Un importe siempre trae 2 decimales; el numero de operacion (17422) no. Ese
# es el discriminador para separar montos de identificadores dentro de la fila.
_RE_BBVA_IMPORTE = re.compile(r'-?[\d,]*\d\.\d{2}(?![\d])')
# Lineas informativas del detalle de consumo en dolares: no son movimientos.
_RE_BBVA_INFORMATIVA = re.compile(r'^\s*IMP\s*\.?\s*OP', re.I)


def _texto_pdf(path, password=None):
    """Texto completo de un PDF, descifrandolo si hace falta. Comparte el
    criterio de parse_eecc_interbank_pdf (ver ahi el porque del descifrado)."""
    import pypdf
    reader = pypdf.PdfReader(path)
    if reader.is_encrypted:
        if not password:
            raise ValueError(
                f'{path}: el PDF esta protegido con contrasena y no se paso ninguna '
                f'(--pdf-password).'
            )
        if reader.decrypt(password) == 0:
            raise ValueError(f'{path}: la contrasena provista no abrio el PDF cifrado.')
    return '\n'.join((p.extract_text() or '') for p in reader.pages)


def _normalizar_importes_pdf(texto):
    """Pega los importes que la extraccion de texto parte antes del punto
    decimal ('4,327 .15' -> '4,327.15', '7 .58' -> '7.58').

    Solo toca digito-espacios-punto-2digitos, asi que no altera 'T .C:' ni
    ninguna abreviatura con punto: ahi lo que precede al espacio es una letra."""
    return re.sub(r'(\d)\s+\.(\d{2})\b', r'\1.\2', texto)


def _filas_logicas_bbva(texto):
    """Agrupa las lineas fisicas en filas logicas. Una fila empieza con dos
    fechas 'DD-MM DD-MM'; lo que viene despues, hasta la siguiente fecha doble,
    pertenece a la misma fila.

    Hace falta porque la extraccion parte una fila en varias lineas cuando la
    descripcion es larga: 'DEPOS. EN CTA. ... OF DOMINGO' / 'CU' /
    'VEN 17425 1,550.00 0.05 5,877.10' son una sola fila, y los importes viven
    en el ultimo pedazo."""
    # El pie del estado ('TOTALES POR ITF' y lo que sigue) trae importes que no
    # son movimientos. Si no se corta ahi, se pegan a la ULTIMA fila y la
    # arruinan: sus importes desplazan al saldo real y la fila entera se
    # descarta. Peor aun, el saldo final se calcula de lo parseado, asi que la
    # fila faltante no descuadra nada y el error pasa desapercibido. Mismo corte
    # que hace parse_eecc_bbva_html.
    corte = re.search(r'TOTALES\s+POR\s+ITF', texto, re.I)
    if corte:
        texto = texto[:corte.start()]

    filas = []
    actual = None
    for linea in texto.split('\n'):
        if _RE_BBVA_INFORMATIVA.match(linea):
            continue
        m = _RE_BBVA_FILA.match(linea.strip())
        if m:
            if actual is not None:
                filas.append(actual)
            actual = [m.group(1), m.group(2), m.group(3).strip()]
        elif actual is not None and linea.strip():
            actual[2] += ' ' + linea.strip()
    if actual is not None:
        filas.append(actual)
    return filas


def parse_eecc_bbva_pdf(path, password=None):
    texto = _normalizar_importes_pdf(_texto_pdf(path, password))

    mes, anio = guess_period_from_filename(path)

    cuenta = ''
    mc = re.search(r'DEPOS\.\s*EN\s*CTA\.\s*(\d{10,})', texto, re.I)
    if mc:
        cuenta = mc.group(1)
    else:
        mfn = re.search(r'_BBVA_(\d+)_', os.path.basename(path), re.I)
        cuenta = mfn.group(1) if mfn else ''

    cliente = ''
    mcli = re.search(r'TITULARES:\s*\n\s*(.+)', texto)
    if mcli:
        cliente = mcli.group(1).strip()

    moneda = 'USD' if re.search(r'MONEDA:\s*D[O\xd3]LARES', texto, re.I) else 'PEN'

    msi = re.search(r'SALDO\s+ANTERIOR\s+(-?[\d,]+\.\d{2})', texto, re.I)
    saldo_inicial = _fnum(msi.group(1)) if msi else None

    movimientos = []
    unparsed = []
    mismatches = []
    anchor_ok = True
    corriente = saldo_inicial

    for fop, fval, resto in _filas_logicas_bbva(texto):
        importes = _RE_BBVA_IMPORTE.findall(resto)
        if len(importes) < 2 or corriente is None:
            unparsed.append(f'{fop} {fval} {resto}')
            continue

        saldo = _fnum(importes[-1])
        elegido = None
        # El saldo del propio documento decide que numero es el monto y cual el
        # ITF: probar la lectura simple primero y quedarse con la que cierra la
        # cadena. Distinguirlos por posicion seria fragil (la columna ITF a
        # veces esta y a veces no).
        candidatos = [(_fnum(importes[-2]), 0.0)]
        if len(importes) >= 3:
            candidatos.append((_fnum(importes[-3]), _fnum(importes[-2]) or 0.0))
        for monto, itf in candidatos:
            if monto is None or saldo is None:
                continue
            if abs(round(corriente + monto - itf, 2) - round(saldo, 2)) <= 0.01:
                elegido = (monto, itf)
                break
        if elegido is None:
            unparsed.append(f'{fop} {fval} {resto}')
            anchor_ok = False
            mismatches.append({'fecha': fop, 'desc': resto[:60],
                                'esperado': None, 'reportado': saldo})
            continue

        monto, itf = elegido
        prefijo = resto[:resto.rfind(importes[-3 if itf else -2])]
        mnro = re.findall(r'\b(\d{4,7})\b', prefijo)
        nro = mnro[-1] if mnro else None
        desc = re.sub(r'\s+', ' ', prefijo.replace(nro, '') if nro else prefijo).strip()

        fop_full = _dashdate_to_ddmmyyyy(fop, mes, anio)
        fval_full = _dashdate_to_ddmmyyyy(fval, mes, anio)
        base = {'fop': fop_full, 'fpr': fval_full, 'nro': nro, 'mov': '',
                'canal': '', '_id': _next_id()}

        saldo_tras_monto = round(corriente + monto, 2)
        movimientos.append(dict(base, desc=desc,
                                 cargo=abs(monto) if monto < 0 else None,
                                 abono=monto if monto >= 0 else None,
                                 saldo=saldo_tras_monto if itf else saldo))
        if itf:
            # Mismo criterio que el parser de BBVA en HTML: el ITF se abre como
            # su propio cargo para que sea visible en la conciliacion.
            movimientos.append({'fop': fop_full, 'fpr': fval_full, 'nro': None,
                                 'mov': '', 'desc': 'ITF', 'canal': '',
                                 'cargo': round(itf, 2), 'abono': None,
                                 'saldo': saldo, '_id': _next_id()})
        corriente = saldo

    saldo_final = movimientos[-1]['saldo'] if movimientos else saldo_inicial

    meta = {
        'banco': 'BBVA', 'formato': 'bbva_pdf', 'cuenta': cuenta, 'moneda': moneda,
        'cliente': cliente, 'periodo_mes': mes, 'periodo_anio': anio,
        'saldo_inicial': saldo_inicial, 'saldo_final': saldo_final,
        'anchor_ok': anchor_ok, 'anchor_mismatches': mismatches,
        'n_movimientos': len(movimientos), 'unparsed_lines': unparsed,
    }
    return movimientos, meta


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------
def parse_eecc(path, password=None):
    """Detecta el formato y devuelve (movimientos, meta). Lanza ValueError con
    mensaje claro si el formato no esta soportado (ej. .xls binario real -
    xlrd esta instalado pero NO sirve para abrir estos archivos, ver SKILL.md).

    password: ver parse_eecc_interbank_pdf. Se ignora para los formatos que
    no son PDF (xlsx, bbva_html no vienen cifrados).

    Un PDF se enruta al parser de su banco mirando el CONTENIDO, no el nombre
    del archivo: el nombre lo pone quien descarga y no es confiable (el EECC de
    BBVA de jul-2026 llego como 'EC_Julio 2026.pdf', sin banco ni cuenta).

    Ademas valida que el resultado no sea un parseo vacio disfrazado de exito:
    ver _validar_parseo."""
    fmt = detect_format(path)
    if fmt == 'xlsx':
        return _validar_parseo(path, *parse_eecc_interbank_xlsx(path))
    if fmt == 'pdf':
        texto = _texto_pdf(path, password)
        if re.search(r'\bBBVA\b', texto):
            return _validar_parseo(path, *parse_eecc_bbva_pdf(path, password=password))
        return _validar_parseo(path, *parse_eecc_interbank_pdf(path, password=password))
    if fmt == 'bbva_html':
        return _validar_parseo(path, *parse_eecc_bbva_html(path))
    if fmt == 'xls_biff_unsupported':
        raise ValueError(
            f'{path}: es un .xls binario (BIFF) real, no un HTML disfrazado de BBVA. '
            'openpyxl no lo puede abrir y xlrd no sirve para este caso (no soporta '
            '.xls modernos/BIFF8 con las hojas que usamos). Reexportar como HTML '
            '(BBVA) o como .xlsx (Interbank), o pasar el PDF oficial.'
        )
    raise ValueError(f'Formato no reconocido para {path}: {fmt}')


def _validar_parseo(path, movimientos, meta):
    """Impide que un parseo fallido se reporte como exito.

    Un EECC que sale con CERO movimientos tiene dos causas muy distintas:

    1. La cuenta de verdad no tuvo movimientos en el periodo (paso en jun-2026
       con la cuenta 4388 de EL TEMPLO). En ese caso el documento SI trae sus
       saldos, y el motor lo reporta como 'SIN MOVIMIENTOS' sin que sea error.
    2. El parser no entendio el documento (formato nuevo, banco equivocado). Ahi
       no hay saldo inicial ni final, porque no se encontro nada. Es el caso que
       hay que gritar: con anchor_ok=True y 0 movimientos, la conciliacion
       reportaria que la cuenta cuadra perfecto mientras ignora el mes entero.
       Paso de verdad con el EECC de BBVA de jul-2026, que llego en PDF y cayo
       en el parser de Interbank.

    El discriminador NO son los saldos: verificado contra el EECC real de la
    4388 de jun-2026, un estado legitimamente vacio tampoco trae saldo inicial
    ni final (no hay ninguna linea de la que sacarlos). Lo que si trae es su
    IDENTIFICACION: cuenta '200-3008494388' y cliente 'EL TEMPLO SAC'. El
    mismo parser sobre un documento que no le corresponde devuelve cuenta y
    cliente vacios, porque no reconocio nada. Esa es la senal fiable: un
    documento leido bien se identifica, aunque no tenga movimientos.

    (La primera version de esta guarda miraba solo los saldos y hacia fallar
    la conciliacion de junio, que es la regresion de referencia del proyecto.)"""
    sin_identificar = not str(meta.get('cuenta') or '').strip() and not str(meta.get('cliente') or '').strip()
    if (not movimientos and sin_identificar
            and meta.get('saldo_inicial') is None and meta.get('saldo_final') is None):
        raise ValueError(
            f'{path}: se parseo a 0 movimientos y sin saldo inicial ni final '
            f'(banco detectado: {meta.get("banco")}, formato: {meta.get("formato")}). '
            f'El documento no corresponde a ese parser o cambio de plantilla. '
            f'Revisar antes de conciliar: un EECC ignorado en silencio deja fuera '
            f'todos los movimientos de esa cuenta.'
        )
    return movimientos, meta
